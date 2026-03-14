#!/usr/bin/env python3
"""
HALO Upload — Push HALO cross-sell campaigns to Amazon Ads via API.

Usage:
  python3 api/halo_upload.py geo                    # upload GEO halo campaigns
  python3 api/halo_upload.py daiken dbj             # upload multiple brands
  python3 api/halo_upload.py flux-de --dry-run      # preview without uploading

Reads from: clients/{brand}/bulk-output/HALO_*.xlsx (latest file)
Creates: SP Campaign → Ad Group → Product Ad → Product Targeting (4 API calls per campaign)
"""

import os
import sys
import json
import time
import csv
from pathlib import Path
from datetime import date

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# ── Load credentials ─────────────────────────────────────
env_path = ROOT / ".env"
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, val = line.split("=", 1)
            if key == "AMAZON_ADS_CLIENT_ID":
                os.environ["AD_API_CLIENT_ID"] = val
            elif key == "AMAZON_ADS_CLIENT_SECRET":
                os.environ["AD_API_CLIENT_SECRET"] = val
            elif key == "AMAZON_ADS_REFRESH_TOKEN":
                os.environ["AD_API_REFRESH_TOKEN"] = val

os.environ["AD_API_PROFILE_ID"] = "0"

from ad_api.api.sp import CampaignsV3, AdGroupsV3, ProductAdsV3, TargetsV3
from ad_api.base.marketplaces import Marketplaces

# ── Brand configs ────────────────────────────────────────
BRANDS = {
    "daiken":   {"profile_id": "1243647931853395", "marketplace": Marketplaces.US, "label": "DAIKEN"},
    "dbj":      {"profile_id": "3565954576012304", "marketplace": Marketplaces.US, "label": "DBJ"},
    "geo":      {"profile_id": "3947387233604911", "marketplace": Marketplaces.US, "label": "GEO"},
    "braintea": {"profile_id": "2695717242300933", "marketplace": Marketplaces.US, "label": "Brain Tea"},
    "flux-au":  {"profile_id": "3601686309797276", "marketplace": Marketplaces.AU, "label": "Flux AU"},
    "flux-uk":  {"profile_id": "2347387401732870", "marketplace": Marketplaces.UK, "label": "Flux UK"},
    "flux-de":  {"profile_id": "4003401656487187", "marketplace": Marketplaces.EU, "label": "Flux DE"},
    "flux-us":  {"profile_id": "3255856020251415", "marketplace": Marketplaces.US, "label": "Flux US"},
    "flux-ca":  {"profile_id": "599373744640869",  "marketplace": Marketplaces.CA, "label": "Flux CA"},
}


def find_latest_halo_bulk(brand_key):
    """Find latest HALO bulk xlsx for a brand."""
    if brand_key.startswith("flux-"):
        bulk_dir = ROOT / "clients/flux/bulk-output"
        prefix = f"HALO_FX-{brand_key.split('-')[1].upper()}_"
    else:
        bulk_dir = ROOT / f"clients/{brand_key}/bulk-output"
        prefix_map = {"daiken": "HALO_DK_", "dbj": "HALO_DBJ_", "geo": "HALO_GEO_", "braintea": "HALO_BT_"}
        prefix = prefix_map.get(brand_key, f"HALO_{brand_key.upper()}_")

    files = sorted(bulk_dir.glob(f"{prefix}*.xlsx"), reverse=True)
    return files[0] if files else None


def parse_bulk(xlsx_path):
    """Parse bulk xlsx into campaign groups (4 rows each)."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    campaigns = []
    current = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        rtype = row[idx["Record Type"]]

        if rtype == "Campaign":
            if current:
                campaigns.append(current)
            current = {
                "name": row[idx["Campaign Name"]],
                "budget": float(row[idx["Campaign Daily Budget"]]),
                "strategy": row[idx["Campaign Bidding Strategy"]],
            }
        elif rtype == "Ad Group":
            current["ag_name"] = row[idx["Ad Group Name"]]
            current["bid"] = float(row[idx["Ad Group Default Bid"]])
        elif rtype == "Product Ad":
            current["sku"] = row[idx["SKU"]]
            current["asin"] = row[idx["ASIN"]]
        elif rtype == "Product Targeting":
            expr = row[idx["Product Targeting Expression"]]
            # Extract target ASIN from asin="B0XXXXXXXX"
            current["target_asin"] = expr.split('"')[1] if '"' in expr else expr
            current["pt_bid"] = float(row[idx["Product Targeting Bid"]])

    if current:
        campaigns.append(current)

    return campaigns


def upload_brand(brand_key, dry_run=False):
    """Upload all HALO campaigns for a brand."""
    cfg = BRANDS[brand_key]
    label = cfg["label"]
    profile_id = cfg["profile_id"]
    marketplace = cfg["marketplace"]

    xlsx_path = find_latest_halo_bulk(brand_key)
    if not xlsx_path:
        print(f"  No HALO bulk file found for {label}")
        return

    campaigns = parse_bulk(xlsx_path)
    print(f"\n{'='*60}")
    print(f"  {label} — {len(campaigns)} HALO campaigns")
    print(f"  File: {xlsx_path.name}")
    print(f"  Profile: {profile_id}")
    if dry_run:
        print(f"  MODE: DRY RUN (no API calls)")
    print(f"{'='*60}")

    if dry_run:
        for c in campaigns:
            print(f"  {c['asin']} → target {c['target_asin']}  bid=${c['bid']}  budget=${c['budget']}")
        print(f"\n  {len(campaigns)} campaigns ready. Remove --dry-run to upload.")
        return

    os.environ["AD_API_PROFILE_ID"] = profile_id
    camp_api = CampaignsV3(marketplace=marketplace)
    ag_api = AdGroupsV3(marketplace=marketplace)
    pa_api = ProductAdsV3(marketplace=marketplace)
    pt_api = TargetsV3(marketplace=marketplace)

    success = 0
    errors = []

    for i, c in enumerate(campaigns, 1):
        print(f"\n  [{i}/{len(campaigns)}] {c['asin']} → {c['target_asin']}")

        # Step 1: Create Campaign
        try:
            resp = camp_api.create_campaigns(
                body={
                    "campaigns": [{
                        "name": c["name"],
                        "state": "ENABLED",
                        "targetingType": "MANUAL",
                        "dynamicBidding": {
                            "strategy": "LEGACY_FOR_SALES",
                        },
                        "budget": {
                            "budgetType": "DAILY",
                            "budget": c["budget"],
                        },
                        "startDate": date.today().isoformat(),
                    }]
                }
            )
            camp_result = resp.payload
            # Response nested: {"campaigns": {"success": [...], "error": [...]}}
            inner = camp_result.get("campaigns", camp_result)
            if inner.get("success") and inner["success"]:
                campaign_id = inner["success"][0]["campaignId"]
                print(f"    Campaign: {campaign_id}")
            else:
                err = inner.get("error", camp_result)
                errors.append(f"{c['name']}: Campaign failed — {err}")
                print(f"    Campaign FAILED: {err}")
                continue
        except Exception as e:
            errors.append(f"{c['name']}: Campaign error — {e}")
            print(f"    Campaign ERROR: {e}")
            continue

        # Step 2: Create Ad Group
        try:
            resp = ag_api.create_ad_groups(
                body={
                    "adGroups": [{
                        "campaignId": campaign_id,
                        "name": c["ag_name"],
                        "state": "ENABLED",
                        "defaultBid": c["bid"],
                    }]
                }
            )
            ag_result = resp.payload
            inner = ag_result.get("adGroups", ag_result)
            if inner.get("success") and inner["success"]:
                ad_group_id = inner["success"][0]["adGroupId"]
                print(f"    Ad Group: {ad_group_id}")
            else:
                err = inner.get("error", ag_result)
                errors.append(f"{c['name']}: Ad Group failed — {err}")
                print(f"    Ad Group FAILED: {err}")
                continue
        except Exception as e:
            errors.append(f"{c['name']}: Ad Group error — {e}")
            print(f"    Ad Group ERROR: {e}")
            continue

        # Step 3: Create Product Ad
        try:
            resp = pa_api.create_product_ads(
                body={
                    "productAds": [{
                        "campaignId": campaign_id,
                        "adGroupId": ad_group_id,
                        "sku": c["sku"],
                        "asin": c["asin"],
                        "state": "ENABLED",
                    }]
                }
            )
            pa_result = resp.payload
            inner = pa_result.get("productAds", pa_result)
            if inner.get("success") and inner["success"]:
                ad_id = inner["success"][0]["adId"]
                print(f"    Product Ad: {ad_id}")
            else:
                err = inner.get("error", pa_result)
                print(f"    Product Ad issue: {err}")
                # Continue anyway — ad might still work
        except Exception as e:
            print(f"    Product Ad ERROR: {e}")

        # Step 4: Create Product Targeting
        try:
            resp = pt_api.create_product_targets(
                body={
                    "targetingClauses": [{
                        "campaignId": campaign_id,
                        "adGroupId": ad_group_id,
                        "state": "ENABLED",
                        "bid": c["pt_bid"],
                        "expression": [
                            {"type": "ASIN_SAME_AS", "value": c["target_asin"]}
                        ],
                        "expressionType": "MANUAL",
                    }]
                }
            )
            pt_result = resp.payload
            inner = pt_result.get("targetingClauses", pt_result)
            if inner.get("success") and inner["success"]:
                target_id = inner["success"][0]["targetId"]
                print(f"    Target: {target_id}")
            else:
                err = inner.get("error", pt_result)
                print(f"    Target issue: {err}")
        except Exception as e:
            print(f"    Target ERROR: {e}")

        success += 1
        time.sleep(0.5)  # Rate limit buffer

    print(f"\n{'─'*60}")
    print(f"  {label}: {success}/{len(campaigns)} campaigns created")
    if errors:
        print(f"  Errors ({len(errors)}):")
        for e in errors:
            print(f"    {e}")
    print(f"{'─'*60}")
    return success


def main():
    raw_args = sys.argv[1:]
    dry_run = "--dry-run" in raw_args
    args = [a for a in raw_args if not a.startswith("--")]

    if not args:
        print(__doc__)
        return

    flux_markets = ["flux-au", "flux-uk", "flux-de", "flux-us", "flux-ca"]
    targets = []
    for a in args:
        if a == "flux":
            targets.extend(flux_markets)
        elif a in BRANDS:
            targets.append(a)
        else:
            print(f"  Unknown brand: {a}")

    for brand_key in targets:
        upload_brand(brand_key, dry_run=dry_run)


if __name__ == "__main__":
    main()
