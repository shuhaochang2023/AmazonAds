"""
Flux DE — Portfolio Structure Excel Generator
Pulls live data from API + Products.csv, generates a clean Excel file.

Usage:
  python3 api/flux_de_portfolio_excel.py
"""

import os
import json
import csv
import re
import time
import requests
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load .env ──
ROOT = Path(__file__).resolve().parent.parent
env = {}
with open(ROOT / '.env') as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            key, val = line.split('=', 1)
            env[key] = val

CLIENT_ID = env['AMAZON_ADS_CLIENT_ID']
CLIENT_SECRET = env['AMAZON_ADS_CLIENT_SECRET']
REFRESH_TOKEN = env['AMAZON_ADS_REFRESH_TOKEN']
DE_PROFILE = '4003401656487187'
API_BASE = 'https://advertising-api-eu.amazon.com'

# ── Portfolio ID map (created 2026-03-13) ──
PORTFOLIO_IDS = {
    '80Days ClassicDB': '215336892338075',
    '80Days ClassicDB2': '90620611616183',
    '80Days LadyGeo': '214256696563679',
    '80Days LadyOval': '23294321520110',
    '80Days PilotDB': '1411123182086',
    '80Days RoundDB': '80133133264840',
    '80Days StrapNylon': '134858511313898',
    '80Days StrapWire': '265573838320487',
    'ICECUBE Avento': '33283584488286',
    'ICECUBE Carmel': '202456044460494',
    'ICECUBE Dynamic': '201331276873687',
    'ICECUBE Hyperbolic': '95321666823369',
    'ICECUBE JetIINeo': '263143800973693',
    'ICECUBE Laso17': '255037387621876',
    'ICECUBE Sportech': '203200250295571',
    'ICECUBE Sprint': '82641404962937',
    'ICECUBE Ventura': '25887230376938',
    'ICECUBE Verano': '201154169219333',
}


def get_token():
    resp = requests.post('https://api.amazon.com/auth/o2/token', data={
        'grant_type': 'refresh_token',
        'refresh_token': REFRESH_TOKEN,
        'client_id': CLIENT_ID,
        'client_secret': CLIENT_SECRET,
    })
    resp.raise_for_status()
    return resp.json()['access_token']


def list_sp_campaigns(token):
    campaigns = []
    headers = {
        'Amazon-Advertising-API-ClientId': CLIENT_ID,
        'Amazon-Advertising-API-Scope': DE_PROFILE,
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/vnd.spCampaign.v3+json',
        'Accept': 'application/vnd.spCampaign.v3+json',
    }
    next_token = None
    while True:
        body = {"maxResults": 100}
        if next_token:
            body["nextToken"] = next_token
        resp = requests.post(f'{API_BASE}/sp/campaigns/list', headers=headers, json=body)
        resp.raise_for_status()
        data = resp.json()
        campaigns.extend(data.get('campaigns', []))
        next_token = data.get('nextToken')
        if not next_token:
            break
        time.sleep(0.3)
    return [c for c in campaigns if c.get('state', '').upper() != 'ARCHIVED']


def load_products():
    products_csv = ROOT / 'clients' / 'flux' / 'input' / 'de' / 'Products.csv'
    child_to_parent = {}
    parent_to_product = {}
    parent_children = defaultdict(list)
    child_info = {}

    with open(products_csv, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            parent = row.get('Parent ASIN', '').strip()
            asin = row.get('ASIN', '').strip()
            title = row.get('Title', '').strip()
            sku = row.get('SKU', '').strip()
            price = row.get('Selling Price', '').strip()
            active = row.get('Active', '').strip()
            if not parent or not asin:
                continue
            t = title.upper()

            # Brand
            if '80DAYS' in t or '80 DAYS' in t or 'BRILLENBAND' in t:
                brand = '80Days'
            else:
                brand = 'ICECUBE'

            # Model
            product = '?'
            for kw, mdl in [('AVENTO','Avento'),('SPORTECH','Sportech'),('VERANO','Verano'),
                            ('VENTURA','Ventura'),('DYNAMIC','Dynamic'),('JET II','JetIINeo'),
                            ('JETII','JetIINeo'),('CARMEL','Carmel'),('HYPERBOLIC','Hyperbolic'),
                            ('SPRINT','Sprint'),('LASO','Laso17'),('PILOTENBRILLE','PilotDB'),
                            ('RUNDE','RoundDB'),('GEOMETRISCH','LadyGeo')]:
                if kw in t:
                    product = mdl
                    break
            if product == '?':
                if 'BRILLENBAND' in t or 'STRAP' in sku.upper():
                    product = 'StrapWire' if '102' in sku else 'StrapNylon'
                for sp, mdl in [('AV101','PilotDB'),('RD101','RoundDB'),('CL102','ClassicDB2'),
                                ('CL101','ClassicDB'),('LD101','LadyGeo'),('LD102','LadyOval')]:
                    if sp in sku:
                        product = mdl
                        break
            if 'ECKIGE' in t and 'CL102' in sku:
                product = 'ClassicDB2'
            if 'OVAL' in t and 'LD102' in sku:
                product = 'LadyOval'

            child_to_parent[asin] = parent
            if asin != parent:
                parent_children[parent].append(asin)
            parent_to_product[parent] = (brand, product)
            child_info[asin] = {
                'sku': sku,
                'title': title[:80] + '...' if len(title) > 80 else title,
                'price': float(price) if price else None,
                'active': active == 'True',
            }

    # Consolidate duplicate Verano
    by_name = defaultdict(list)
    for p, (b, m) in parent_to_product.items():
        by_name[f"{b} {m}"].append(p)
    for name, parents in by_name.items():
        if len(parents) > 1:
            best = max(parents, key=lambda p: len([c for c in parent_children.get(p,[]) if child_info.get(c,{}).get('active')]))
            for other in parents:
                if other != best:
                    for child in parent_children.get(other, []):
                        child_to_parent[child] = best
                        parent_children[best].append(child)
                    del parent_to_product[other]
                    if other in parent_children:
                        del parent_children[other]

    return child_to_parent, parent_to_product, parent_children, child_info


ASIN_RE = re.compile(r'(B0[A-Z0-9]{8,10})')

NAME_HEURISTICS = [
    (r'(?i)(?:^|[\s_-])strap', '80Days', 'StrapNylon'),
    (r'(?i)avento', 'ICECUBE', 'Avento'),
    (r'(?i)sportech', 'ICECUBE', 'Sportech'),
    (r'(?i)verano', 'ICECUBE', 'Verano'),
    (r'(?i)ventura', 'ICECUBE', 'Ventura'),
    (r'(?i)dynamic', 'ICECUBE', 'Dynamic'),
    (r'(?i)jet.?ii|jetii', 'ICECUBE', 'JetIINeo'),
    (r'(?i)carmel', 'ICECUBE', 'Carmel'),
    (r'(?i)hyperbolic', 'ICECUBE', 'Hyperbolic'),
    (r'(?i)sprint', 'ICECUBE', 'Sprint'),
    (r'(?i)laso', 'ICECUBE', 'Laso17'),
    (r'(?i)pilot', '80Days', 'PilotDB'),
    (r'(?i)round', '80Days', 'RoundDB'),
    (r'(?i)classic', '80Days', 'ClassicDB'),
    (r'(?i)LD101', '80Days', 'LadyGeo'),
    (r'(?i)LD102', '80Days', 'LadyOval'),
    (r'(?i)ICECUBE.+All', 'ICECUBE', 'Avento'),
]


def parse_match_type(name):
    n = name.lower()
    if 'auto' in n: return 'Auto'
    if 'exact' in n: return 'KW Exact'
    if 'phrase' in n: return 'KW Phrase'
    if 'broad' in n: return 'KW Broad'
    if 'product target' in n or '_pt_' in n or ' pt ' in n or ' pt-' in n: return 'Product Targeting'
    if 'category' in n or '_cat_' in n: return 'Category'
    if ' asin ' in n: return 'ASIN Targeting'
    if ' kw ' in n: return 'KW'
    return 'Other'


def main():
    now = datetime.now()
    print("=" * 60)
    print("  FLUX DE — Portfolio Structure Excel Generator")
    print("=" * 60)

    child_to_parent, parent_to_product, parent_children, child_info = load_products()
    print(f"  {len(parent_to_product)} parents, {len(child_to_parent)} ASINs")

    token = get_token()
    sp_campaigns = list_sp_campaigns(token)
    print(f"  {len(sp_campaigns)} SP campaigns (non-archived)")

    # Match campaigns
    campaign_data = []
    unmatched = []

    for c in sp_campaigns:
        cid = c.get('campaignId', '')
        name = c.get('name', '')
        state = c.get('state', '')
        budget = c.get('budget', {})
        daily_budget = budget.get('budget', 0) if isinstance(budget, dict) else 0
        start_date = c.get('startDate', '')
        portfolio_id = c.get('portfolioId', '')

        parent = None
        child_asin = None
        for asin in ASIN_RE.findall(name):
            if asin in child_to_parent:
                parent = child_to_parent[asin]
                child_asin = asin
                break

        brand, product = '?', '?'
        if parent and parent in parent_to_product:
            brand, product = parent_to_product[parent]
        elif not parent:
            for pattern, b, p in NAME_HEURISTICS:
                if re.search(pattern, name):
                    brand, product = b, p
                    # Find parent for this brand/product
                    for pp, (pb, pm) in parent_to_product.items():
                        if pb == b and pm == p:
                            parent = pp
                            break
                    break

        entry = {
            'type': 'SP',
            'campaignId': cid,
            'name': name,
            'state': state,
            'dailyBudget': daily_budget,
            'startDate': start_date,
            'portfolioId': portfolio_id,
            'parent': parent,
            'childAsin': child_asin,
            'brand': brand,
            'product': product,
            'portfolio': f'{brand} {product}',
            'matchType': parse_match_type(name),
        }
        if parent:
            campaign_data.append(entry)
        else:
            unmatched.append(entry)

    print(f"  Matched: {len(campaign_data)} | Unmatched: {len(unmatched)}")

    # ── BUILD EXCEL ──
    wb = openpyxl.Workbook()
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    subtitle_font = Font(name='Calibri', size=10, color='666666')
    thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))
    wrap = Alignment(wrap_text=True, vertical='top')

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sort parents
    all_parents = sorted(parent_to_product.keys(), key=lambda p: (
        parent_to_product[p][0], parent_to_product[p][1]
    ))

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = 'Overview'
    ws.cell(1, 1, 'Flux DE — Portfolio Structure Overview').font = title_font
    ws.cell(2, 1, f'Generated: {now.strftime("%Y-%m-%d %H:%M")}  |  Profile: {DE_PROFILE}').font = subtitle_font

    hdrs = ['Portfolio', 'Brand', 'Product', 'Parent ASIN', 'Children (CSV)', 'SP Campaigns', 'Enabled', 'Paused']
    for i, h in enumerate(hdrs, 1):
        ws.cell(4, i, h)
    style_header(ws, 4, len(hdrs))

    row = 5
    t_ch = t_sp = t_en = t_pa = 0
    for parent in all_parents:
        brand, product = parent_to_product[parent]
        portfolio = f'{brand} {product}'
        children = [c for c in parent_children.get(parent, []) if c != parent]
        sp = sum(1 for c in campaign_data if c['parent'] == parent)
        en = sum(1 for c in campaign_data if c['parent'] == parent and c['state'].upper() == 'ENABLED')
        pa = sum(1 for c in campaign_data if c['parent'] == parent and c['state'].upper() == 'PAUSED')
        ws.cell(row, 1, portfolio); ws.cell(row, 2, brand); ws.cell(row, 3, product)
        ws.cell(row, 4, parent); ws.cell(row, 5, len(children))
        ws.cell(row, 6, sp); ws.cell(row, 7, en); ws.cell(row, 8, pa)
        for col in range(1, len(hdrs)+1): ws.cell(row, col).border = thin_border
        t_ch += len(children); t_sp += sp; t_en += en; t_pa += pa
        row += 1
    ws.cell(row, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(row, 5, t_ch).font = Font(bold=True)
    ws.cell(row, 6, t_sp).font = Font(bold=True)
    ws.cell(row, 7, t_en).font = Font(bold=True)
    ws.cell(row, 8, t_pa).font = Font(bold=True)
    row += 2
    ws.cell(row, 1, f'Unmatched campaigns: {len(unmatched)}').font = subtitle_font
    for col, w in enumerate([22, 10, 14, 16, 14, 14, 10, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Product Catalog ──
    ws2 = wb.create_sheet('Product Catalog')
    ws2.cell(1, 1, 'Product Catalog — Parent → Child ASIN Mapping (DE)').font = title_font
    hdrs2 = ['Portfolio', 'Brand', 'Product', 'Parent ASIN', 'Child ASIN', 'SKU', 'Title', 'Price (EUR)', 'Active']
    for i, h in enumerate(hdrs2, 1): ws2.cell(3, i, h)
    style_header(ws2, 3, len(hdrs2))

    row = 4
    for parent in all_parents:
        brand, product = parent_to_product[parent]
        portfolio = f'{brand} {product}'
        children = [c for c in parent_children.get(parent, []) if c != parent]
        ws2.cell(row, 1, portfolio); ws2.cell(row, 2, brand); ws2.cell(row, 3, product)
        ws2.cell(row, 4, parent); ws2.cell(row, 4).font = Font(bold=True)
        row += 1
        for child in sorted(children):
            info = child_info.get(child, {})
            ws2.cell(row, 5, child); ws2.cell(row, 6, info.get('sku', ''))
            ws2.cell(row, 7, info.get('title', '')); ws2.cell(row, 7).alignment = wrap
            ws2.cell(row, 8, info.get('price')); ws2.cell(row, 9, 'Yes' if info.get('active') else 'No')
            if not info.get('active'):
                for col in range(5, 10): ws2.cell(row, col).font = Font(color='999999')
            for col in range(1, len(hdrs2)+1): ws2.cell(row, col).border = thin_border
            row += 1
        row += 1
    for col, w in enumerate([22, 10, 14, 16, 16, 22, 55, 12, 8], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Targeting Mix ──
    ws3 = wb.create_sheet('Targeting Mix')
    ws3.cell(1, 1, 'Targeting Type Breakdown by Product').font = title_font
    match_types = ['Auto', 'KW', 'KW Broad', 'KW Exact', 'KW Phrase', 'ASIN Targeting', 'Product Targeting', 'Category', 'Other']
    hdrs3 = ['Product', 'Brand', 'Portfolio', 'Total'] + match_types
    for i, h in enumerate(hdrs3, 1): ws3.cell(3, i, h)
    style_header(ws3, 3, len(hdrs3))

    row = 4
    for parent in all_parents:
        brand, product = parent_to_product[parent]
        portfolio = f'{brand} {product}'
        pc = [c for c in campaign_data if c['parent'] == parent]
        ws3.cell(row, 1, product); ws3.cell(row, 2, brand); ws3.cell(row, 3, portfolio)
        ws3.cell(row, 4, len(pc))
        for j, mt in enumerate(match_types, 5):
            cnt = sum(1 for c in pc if c['matchType'] == mt)
            ws3.cell(row, j, cnt if cnt > 0 else None)
        for col in range(1, len(hdrs3)+1): ws3.cell(row, col).border = thin_border
        row += 1
    for col, w in enumerate([14, 10, 22, 8] + [10]*len(match_types), 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Dashboard Verify ──
    ws4 = wb.create_sheet('Dashboard Verify')
    ws4.cell(1, 1, 'Dashboard Parent-Child Verification (DE)').font = title_font
    ws4.cell(2, 1, 'Compare against de.html dashboard data').font = subtitle_font
    hdrs4 = ['Parent ASIN', 'Product', 'Brand', 'Portfolio', 'Children (CSV)', 'Children ASINs', 'SP Campaigns', 'Active Children', 'Inactive Children']
    for i, h in enumerate(hdrs4, 1): ws4.cell(4, i, h)
    style_header(ws4, 4, len(hdrs4))

    row = 5
    for parent in all_parents:
        brand, product = parent_to_product[parent]
        portfolio = f'{brand} {product}'
        children = [c for c in parent_children.get(parent, []) if c != parent]
        active_ch = [c for c in children if child_info.get(c, {}).get('active')]
        inactive_ch = [c for c in children if not child_info.get(c, {}).get('active')]
        sp = sum(1 for c in campaign_data if c['parent'] == parent)
        ws4.cell(row, 1, parent); ws4.cell(row, 2, product); ws4.cell(row, 3, brand)
        ws4.cell(row, 4, portfolio); ws4.cell(row, 5, len(children))
        ws4.cell(row, 6, ', '.join(sorted(children))); ws4.cell(row, 6).alignment = wrap
        ws4.cell(row, 7, sp); ws4.cell(row, 8, len(active_ch)); ws4.cell(row, 9, len(inactive_ch))
        for col in range(1, len(hdrs4)+1): ws4.cell(row, col).border = thin_border
        row += 1
    for col, w in enumerate([16, 14, 10, 22, 14, 55, 14, 14, 14], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ── Save ──
    filename = f'20260313_Flux_DE_Portfolio_Structure.xlsx'
    out_path = ROOT / 'flux' / filename
    wb.save(out_path)
    print(f"\n  Saved: {out_path}")
    print(f"  {len(all_parents)} parents | {t_ch} children | {len(campaign_data)} matched | {len(unmatched)} unmatched")


if __name__ == '__main__':
    main()
