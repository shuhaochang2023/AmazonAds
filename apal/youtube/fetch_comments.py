#!/usr/bin/env python3
"""
APAL YouTube Comment Fetcher & Categorizer
==========================================
Fetches all comments from a YouTube video, categorizes them,
and exports to Excel.

Setup:
  1. Go to https://console.cloud.google.com/
  2. Create a project (or use existing)
  3. Enable "YouTube Data API v3"
  4. Create credentials:
     - For READ-ONLY (fetch comments): API Key is enough
     - For REPLY (post comments): OAuth 2.0 Client ID (Desktop app)
  5. Save:
     - API Key → paste below or set env var YOUTUBE_API_KEY
     - OAuth JSON → save as client_secret.json in this folder

Usage:
  # Fetch & categorize comments
  python3 fetch_comments.py VIDEO_ID

  # Or with full URL
  python3 fetch_comments.py "https://www.youtube.com/watch?v=X4q7l_8pQHg"

  # Compare with previous export to find NEW comments only
  python3 fetch_comments.py VIDEO_ID --diff previous_export.xlsx
"""

import sys
import os
import re
import json
import argparse
from datetime import datetime

from googleapiclient.discovery import build

# ─── CONFIG ───
API_KEY = os.environ.get("YOUTUBE_API_KEY", "")  # paste your key here or set env var
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ─── CATEGORIZATION RULES ───
def categorize(text):
    """Categorize a comment and return (category, intent)."""
    txt = text.lower().strip()

    # --- CORE STRENGTHS ---
    strength_phrases = [
        'love', 'awesome', 'amazing', 'excellent', 'fantastic', 'brilliant',
        'game changer', 'game-changer', 'high quality', 'well made', 'well-made',
        'impressed', 'impressive', 'incredible', 'wonderful', 'finally',
        'quality unit', 'solid build', 'delightful', 'authentic',
    ]
    # Exclude if also mentions framerate/fps issues
    if any(p in txt for p in strength_phrases) and not any(p in txt for p in ['frame', 'fps', 'potato']):
        if 'learn' in txt:
            return "Core Strengths", "Values educational/discovery format"
        if 'authentic' in txt:
            return "Core Strengths", "Values authenticity — genuine review resonates"
        if 'quality' in txt or 'well made' in txt or 'solid' in txt:
            return "Core Strengths", "Product quality validation"
        return "Core Strengths", "Positive brand/product sentiment"

    # --- PAIN POINTS ---
    pain_phrases = [
        'too expensive', 'overpriced', 'confusing', 'broke', 'broken',
        'hard to', 'difficult', 'frustrat', 'disappoint', 'doesn\'t work',
        'not work', 'poor', 'bad quality', '30kb', '30 kb', 'data limit',
        'limited data', 'last commit', 'abandoned', 'dead project',
        'no documentation', 'no docs', 'stale', 'years ago',
        'what can it do', 'of what use', 'doesn\'t even',
    ]
    if any(p in txt for p in pain_phrases):
        if '30kb' in txt or '30 kb' in txt or 'data limit' in txt or 'airtime' in txt:
            return "Pain Points", "Data limit concern — 30KB feels impractical; needs use-case framing"
        if 'commit' in txt or 'years ago' in txt or 'abandoned' in txt:
            return "Pain Points", "Software maintenance concern — stale repos signal abandonment risk"
        if 'expensive' in txt or 'overpriced' in txt:
            return "Pain Points", "Price/value concern"
        if 'what can it do' in txt or 'of what use' in txt:
            return "Pain Points", "Value proposition unclear"
        return "Pain Points", "Product issue or criticism"

    # --- BUSINESS OPPORTUNITIES ---
    biz_phrases = [
        'where can i buy', 'how can i get', 'where did you get',
        'get my hands on', 'i want one', "i'd like to", 'how much',
        'is a sim card', 'available', 'purchase', 'dm me',
        'interested in', 'experiment with', 'order one',
    ]
    usecase_phrases = [
        'sensor', 'weather station', 'building automation', 'fleet',
        'scada', 'wastewater', 'ship', 'marine', 'boat', 'sailing',
        'off-grid', 'remote monitor', 'emergency', 'camping',
        'farm', 'ranch', 'fire detection', 'movement sensor',
        'tracking', 'modbus', 'plc', 'bitcoin', 'wallet',
        'mesh to sat', 'subscription', 'overseas', 'contract',
        'iot', 'telemetry', 'security system', 'radiation',
        'm.2 card', 'gunshot',
    ]
    if any(p in txt for p in biz_phrases):
        if any(p in txt for p in ['buy', 'get', 'hands on', 'want one', 'how much', 'purchase', 'order']):
            return "Business Opportunities", "Purchase intent — high-intent buyer"
        if 'dm me' in txt:
            return "Business Opportunities", "Direct sales lead / B2B inquiry"
        return "Business Opportunities", "Product interest / availability question"
    if any(p in txt for p in usecase_phrases):
        if 'modbus' in txt or 'plc' in txt or 'scada' in txt:
            return "Business Opportunities", "Industrial IoT integration opportunity"
        if any(p in txt for p in ['ship', 'marine', 'boat', 'sailing']):
            return "Business Opportunities", "Marine/maritime use case"
        if any(p in txt for p in ['farm', 'ranch', 'off-grid', 'remote']):
            return "Business Opportunities", "Remote/off-grid use case"
        return "Business Opportunities", "Use case expansion — community-suggested application"

    # --- COMPETITOR COMPARISONS ---
    comp_phrases = [
        'meshtastic', 'meshcore', 'reticulum', 'winlink',
        'lorawan', 'lora ', 'lora.', 'starlink', 'iridium',
        'garmin', 'inreach', 'globalstar', 'swarm', 'helium',
        'inmarsat', 'nodered', 'node-red', 'node red',
        'raspberry pi', 'esp32', 'adafruit', 'compared to',
        'better than', 'cheaper than', 'zoleo',
    ]
    if any(p in txt for p in comp_phrases):
        if 'meshtastic' in txt or 'meshcore' in txt or 'reticulum' in txt or 'winlink' in txt:
            return "Competitor Comparisons", "Mesh protocol ecosystem — interoperability interest"
        if 'lorawan' in txt or 'lora ' in txt:
            return "Competitor Comparisons", "LoRa/LoRaWAN ecosystem comparison"
        if 'starlink' in txt:
            return "Competitor Comparisons", "Starlink comparison — different use case (broadband vs IoT)"
        if 'inmarsat' in txt:
            return "Competitor Comparisons", "Legacy satellite comparison"
        return "Competitor Comparisons", "Platform/ecosystem comparison"

    # --- NOISE ---
    return "Noise", ""


def extract_video_id(url_or_id):
    """Extract video ID from URL or return as-is."""
    m = re.search(r'(?:v=|youtu\.be/|/embed/|/v/)([A-Za-z0-9_-]{11})', url_or_id)
    return m.group(1) if m else url_or_id


def fetch_all_comments(youtube, video_id):
    """Fetch all top-level comments + replies for a video."""
    comments = []
    request = youtube.commentThreads().list(
        part="snippet,replies",
        videoId=video_id,
        maxResults=100,
        textFormat="plainText",
        order="time"
    )

    page = 0
    while request:
        page += 1
        print(f"  Fetching page {page}...")
        response = request.execute()

        for item in response.get("items", []):
            # Top-level comment
            top = item["snippet"]["topLevelComment"]["snippet"]
            comments.append({
                "comment_id": item["snippet"]["topLevelComment"]["id"],
                "author": top["authorDisplayName"],
                "comment": top["textDisplay"],
                "likes": top["likeCount"],
                "published": top["publishedAt"],
                "updated": top["updatedAt"],
                "is_reply": False,
                "parent_id": None,
            })

            # Replies
            if item.get("replies"):
                for reply in item["replies"]["comments"]:
                    r = reply["snippet"]
                    comments.append({
                        "comment_id": reply["id"],
                        "author": r["authorDisplayName"],
                        "comment": r["textDisplay"],
                        "likes": r["likeCount"],
                        "published": r["publishedAt"],
                        "updated": r["updatedAt"],
                        "is_reply": True,
                        "parent_id": item["snippet"]["topLevelComment"]["id"],
                    })

        request = youtube.commentThreads().list_next(request, response)

    return comments


def export_excel(comments, video_id, diff_file=None):
    """Export categorized comments to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Load previous comments for diff
    old_authors_comments = set()
    if diff_file and os.path.exists(diff_file):
        old_wb = openpyxl.load_workbook(diff_file)
        for sheet in old_wb.sheetnames:
            ws = old_wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    old_authors_comments.add((str(row[0]).strip(), str(row[1]).strip()[:80]))
        print(f"  Loaded {len(old_authors_comments)} previous comments for diff")

    # Categorize
    for c in comments:
        cat, intent = categorize(c['comment'])
        c['category'] = cat
        c['intent'] = intent
        c['is_new'] = (c['author'].strip(), c['comment'].strip()[:80]) not in old_authors_comments if old_authors_comments else True

    new_count = sum(1 for c in comments if c['is_new'])
    print(f"\n  Total: {len(comments)} | New: {new_count}")

    # Colors
    COLORS = {
        'Core Strengths': ('16a34a', 'dcfce7'),
        'Pain Points': ('dc2626', 'fee2e2'),
        'Business Opportunities': ('2563eb', 'dbeafe'),
        'Competitor Comparisons': ('f59e0b', 'fef3c7'),
        'Noise': ('94a3b8', 'f1f5f9'),
    }

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1e293b')
    wrap = Alignment(wrap_text=True, vertical='top')
    thin = Border(*[Side(style='thin', color='e2e8f0')] * 4)

    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = 'Summary'
    from collections import Counter
    cats = Counter(c['category'] for c in comments)
    cats_new = Counter(c['category'] for c in comments if c['is_new'])

    ws.append(['Category', 'Total', 'New', 'Key Insight'])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    insights = {
        'Core Strengths': 'Positive sentiment — double down on these in marketing',
        'Business Opportunities': 'Purchase intent + use cases — fix sales funnel, highlight applications',
        'Competitor Comparisons': 'Ecosystem context — clarify differentiation and compatibility',
        'Pain Points': 'Actionable feedback — address in product/marketing/docs',
        'Noise': 'Off-topic or non-actionable',
    }
    for cat in ['Core Strengths', 'Business Opportunities', 'Competitor Comparisons', 'Pain Points', 'Noise']:
        ws.append([cat, cats.get(cat, 0), cats_new.get(cat, 0), insights[cat]])
        row = ws.max_row
        fg, bg = COLORS[cat]
        ws.cell(row, 1).fill = PatternFill('solid', fgColor=bg)
        ws.cell(row, 1).font = Font(bold=True, color=fg)

    ws.append([])
    ws.append(['Total', len(comments), new_count])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 70

    # --- Category sheets ---
    for cat in ['Core Strengths', 'Business Opportunities', 'Competitor Comparisons', 'Pain Points']:
        safe = cat.replace(' ', '_')[:31]
        ws = wb.create_sheet(safe)
        ws.append(['NEW?', 'Author', 'Comment', 'Intent', 'Likes', 'Date', 'Reply?', 'Comment ID'])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill('solid', fgColor=COLORS[cat][0])

        for c in sorted([x for x in comments if x['category'] == cat], key=lambda x: x['published']):
            ws.append([
                'NEW' if c['is_new'] else '',
                c['author'], c['comment'], c['intent'],
                c['likes'], c['published'][:10],
                'Reply' if c['is_reply'] else 'Top',
                c['comment_id']
            ])
            row = ws.max_row
            if c['is_new']:
                ws.cell(row, 1).fill = PatternFill('solid', fgColor='fef08a')
                ws.cell(row, 1).font = Font(bold=True, color='a16207')
            for col in range(1, 9):
                ws.cell(row, col).alignment = wrap

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 6
        ws.column_dimensions['H'].width = 30

    # --- All comments ---
    ws = wb.create_sheet('All_Comments')
    ws.append(['NEW?', 'Author', 'Comment', 'Category', 'Intent', 'Likes', 'Date', 'Comment ID'])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in sorted(comments, key=lambda x: x['published']):
        ws.append([
            'NEW' if c['is_new'] else '',
            c['author'], c['comment'], c['category'], c['intent'],
            c['likes'], c['published'][:10], c['comment_id']
        ])
        row = ws.max_row
        if c['category'] in COLORS:
            fg, bg = COLORS[c['category']]
            ws.cell(row, 4).fill = PatternFill('solid', fgColor=bg)
            ws.cell(row, 4).font = Font(color=fg, bold=True)
        if c['is_new']:
            ws.cell(row, 1).fill = PatternFill('solid', fgColor='fef08a')
        for col in range(1, 9):
            ws.cell(row, col).alignment = wrap

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 30

    # Save
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f"APAL_Comments_{video_id}_{ts}.xlsx"
    outpath = os.path.join(OUTPUT_DIR, filename)
    wb.save(outpath)
    print(f"\n  Saved: {outpath}")
    return outpath


def main():
    parser = argparse.ArgumentParser(description="Fetch & categorize YouTube comments")
    parser.add_argument("video", help="YouTube video URL or ID")
    parser.add_argument("--diff", help="Previous Excel to compare for NEW comments", default=None)
    parser.add_argument("--key", help="YouTube API key (or set YOUTUBE_API_KEY env var)", default=None)
    parser.add_argument("--json", help="Also save raw JSON", action="store_true")
    args = parser.parse_args()

    api_key = args.key or API_KEY
    if not api_key:
        print("ERROR: No API key provided.")
        print("  Option 1: python3 fetch_comments.py VIDEO_ID --key YOUR_KEY")
        print("  Option 2: export YOUTUBE_API_KEY=YOUR_KEY")
        print("  Option 3: Edit API_KEY in this script")
        print("\nTo get an API key:")
        print("  1. Go to https://console.cloud.google.com/")
        print("  2. Create/select a project")
        print("  3. Enable 'YouTube Data API v3'")
        print("  4. Go to Credentials → Create → API Key")
        sys.exit(1)

    video_id = extract_video_id(args.video)
    print(f"Video ID: {video_id}")

    youtube = build("youtube", "v3", developerKey=api_key)

    # Fetch video title
    vid_resp = youtube.videos().list(part="snippet", id=video_id).execute()
    if vid_resp["items"]:
        title = vid_resp["items"][0]["snippet"]["title"]
        print(f"Title: {title}")

    print(f"\nFetching comments...")
    comments = fetch_all_comments(youtube, video_id)
    print(f"  Fetched {len(comments)} comments ({sum(1 for c in comments if c['is_reply'])} replies)")

    if args.json:
        json_path = os.path.join(OUTPUT_DIR, f"comments_{video_id}.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(comments, f, ensure_ascii=False, indent=2)
        print(f"  JSON saved: {json_path}")

    export_excel(comments, video_id, args.diff)


if __name__ == "__main__":
    main()
