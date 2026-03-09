#!/usr/bin/env python3
"""
UK Ad Spend Rebalance — Pause bleeders, reduce over-spenders, grow good performers.

Uses Campaign IDs from UK_bulk.xlsx (downloaded 3/9).
"""

import openpyxl
from openpyxl.utils import get_column_letter

TIMESTAMP = '20260310_1500'

SP_HEADERS = [
    'Product', 'Entity', 'Operation', 'Campaign ID', 'Ad Group ID',
    'Portfolio ID', 'Ad ID', 'Keyword ID', 'Product Targeting ID',
    'Campaign Name', 'Ad Group Name',
    'Campaign Name (Informational only)', 'Ad Group Name (Informational only)',
    'Portfolio Name (Informational only)', 'Start Date', 'End Date',
    'Targeting Type', 'State',
    'Campaign State (Informational only)', 'Ad Group State (Informational only)',
    'Daily Budget', 'SKU', 'ASIN (Informational only)',
    'Eligibility Status (Informational only)',
    'Reason for Ineligibility (Informational only)',
    'Ad Group Default Bid', 'Ad Group Default Bid (Informational only)', 'Bid',
    'Keyword Text', 'Native Language Keyword', 'Native Language Locale',
    'Match Type', 'Bidding Strategy', 'Placement', 'Percentage',
    'Product Targeting Expression',
    'Resolved Product Targeting Expression (Informational only)',
    'Audience ID', 'Shopper Cohort Percentage', 'Shopper Cohort Type',
    'Segment Name (Informational only)',
    'Impressions', 'Clicks', 'Click-through Rate', 'Spend', 'Sales',
    'Orders', 'Units', 'Conversion Rate', 'ACOS', 'CPC', 'ROAS',
]

SP_ID_COLS = [3, 4, 5, 6, 7, 8]


def set_id_format(ws, id_col_indices, max_row):
    for col_idx in id_col_indices:
        col_letter = get_column_letter(col_idx + 1)
        for row in range(1, max_row + 1):
            ws[f'{col_letter}{row}'].number_format = '@'


def write_sheet(wb, sheet_name, headers, rows, id_col_indices):
    ws = wb.active
    ws.title = sheet_name
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=val)
    set_id_format(ws, id_col_indices, len(rows) + 1)


def pause_campaign(campaign_id):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Campaign'
    row['Operation'] = 'Update'
    row['Campaign ID'] = str(campaign_id)
    row['State'] = 'paused'
    return row


def make_campaign_row(campaign_name, daily_budget, targeting_type='Auto'):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Campaign'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Campaign Name'] = campaign_name
    row['State'] = 'enabled'
    row['Daily Budget'] = daily_budget
    row['Targeting Type'] = targeting_type
    row['Bidding Strategy'] = 'Dynamic bids - down only'
    return row


def make_ad_group_row(campaign_name, ad_group_name, default_bid):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Ad Group'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['Ad Group Name'] = ad_group_name
    row['State'] = 'enabled'
    row['Ad Group Default Bid'] = default_bid
    return row


def make_product_ad_row(campaign_name, ad_group_name, sku):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Product Ad'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['SKU'] = sku
    row['State'] = 'enabled'
    return row


def make_keyword_row(campaign_name, ad_group_name, keyword, match_type, bid):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Keyword'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['Keyword Text'] = keyword
    row['Match Type'] = match_type
    row['Bid'] = bid
    row['State'] = 'enabled'
    return row


rows = []

# ═══════════════════════════════════════════════════════════════
# SECTION A: PAUSE BLEEDERS (113 campaigns → paused)
# ═══════════════════════════════════════════════════════════════

# --- 80Days Pilot / Women's Sunglasses B0F1CZQRYS --- (51 campaigns, 332.7% ACOS)
pilot_ids = [
    '282420968661693', '507601886205050', '463749428479346', '364534042395812',
    '531518989140668', '308904559326050', '451200101313386', '341590543372331',
    '559649071920906', '494620862518986', '386085275012446', '373388820967280',
    '503981435908050', '313889290029201', '541315071757231', '423738841940681',
    '320964930046624', '346288100262117', '309179546052853', '457971974497951',
    '533250678822141', '301816079511692', '519154999404567', '445164998584386',
    '487284369300720', '535026961592557', '491611681658943', '350293387215993',
    '395388437150262', '297794657216446', '482577713094112', '333615408107052',
    '385540273543014', '309495146791733', '437693328787723', '512994915843485',
    '333229833801656', '465661472046294', '448919084393791', '315643314831102',
    '436843042362257', '497111800175937', '389878273784258', '431445343533531',
    '518113175237308', '407251563758323', '347005652327905', '8353216191360',
    '70935087793019', '109529490088653', '260431061034377',
]

# --- 80Days Oval B0F1D2NSWN --- (16 campaigns, 0 sales)
oval_ids = [
    '290836289165648', '375036152131749', '509598143146341', '314097368322209',
    '307224428423799', '506113572247122', '413388867530597', '550836418373368',
    '458058020482685', '380745635979140', '461658979055814', '556916182926270',
    '313438088556579', '417933474032103', '381914267810044', '560793704918965',
]

# --- 80Days Round B0F1CT/CV --- (46 campaigns, 0 sales)
round_ids = [
    '71183403345235', '237071151371001', '228746178211840', '9556685504309',
    '255197227294288', '144933353580908', '142327712533160', '131113986515427',
    '128860339333451', '244453728479147', '130657404897044', '161963080684398',
    '151956231153853', '211123003339513', '192587496438193', '129757979332348',
    '36920083708778', '5810577375148', '248834489740984', '97948549899866',
    '47792458437295', '243269475757252', '320353143496278', '193734449741092',
    '130116254318620', '253992548524245', '44988923920497', '70365208968781',
    '261191192781279', '79559047762370', '156675181250888', '233084714425379',
    '146121855437689', '273529271533973', '38524088628741', '265016718204008',
    '185964800916791', '75846863457665', '219750556960276', '1488575124058',
    '16427386644649', '118241300898704', '183775460760471', '97692423315538',
    '137396978379372', '85295531492896',
]

print('=== SECTION A: PAUSE BLEEDERS ===')
print(f'  Pilot/Women Sunglasses: {len(pilot_ids)} campaigns (332.7% ACOS)')
print(f'  Oval Sunglasses: {len(oval_ids)} campaigns (0 sales)')
print(f'  Round Sunglasses: {len(round_ids)} campaigns (0 sales)')

for cid in pilot_ids + oval_ids + round_ids:
    rows.append(pause_campaign(cid))

# Also pause zero-sale Sportech campaigns (high ACOS overall 113.5%)
# Only pause those with spend > 0 and zero sales
sportech_pause_ids = [
    '498210793745078',  # £4.90 spend, 0 sales
    '327308238407476',  # £3.38 spend, 0 sales
    '328239638161076',  # £2.97 spend, 0 sales
    '515766331958579',  # £2.75 spend, 0 sales
    '487659834999280',  # £2.21 spend, 0 sales
    '294136432721771',  # £2.17 spend, 0 sales
    '559791901711228',  # £1.91 spend, 0 sales
    '474824686069339',  # £1.87 spend, 0 sales
    '381970222613920',  # £1.84 spend, 0 sales
    '433893466253194',  # £1.76 spend, 0 sales (Verano ASIN in Sportech campaign)
    '546544520042231',  # £1.70 spend, 0 sales
    '343743896632238',  # £1.68 spend, 0 sales
    '430316994237150',  # £1.68 spend, 0 sales
]
print(f'  Sportech zero-sale: {len(sportech_pause_ids)} campaigns')

for cid in sportech_pause_ids:
    rows.append(pause_campaign(cid))

# Pause zero-sale Nylon Strap campaigns
strap_pause_ids = [
    '421977427943212',  # £3.31 spend, 0 sales
    '454322582321263',  # £3.20 spend, 0 sales
    '452427078964025',  # £2.36 spend, 0 sales
    '523213776924993',  # £1.54 spend, 0 sales
    '534763129984377',  # £1.50 spend, 0 sales
    '300783163634464',  # £1.06 spend, 0 sales
    '538952337610754',  # £0.76 spend, 0 sales
]
print(f'  Strap zero-sale: {len(strap_pause_ids)} campaigns')

for cid in strap_pause_ids:
    rows.append(pause_campaign(cid))

total_pause = len(pilot_ids) + len(oval_ids) + len(round_ids) + len(sportech_pause_ids) + len(strap_pause_ids)
print(f'  TOTAL PAUSED: {total_pause} campaigns')

# ═══════════════════════════════════════════════════════════════
# SECTION B: NEW GROWTH CAMPAIGNS
# ═══════════════════════════════════════════════════════════════

print('\n=== SECTION B: NEW GROWTH CAMPAIGNS ===')

# 1. Avento Exact Match (high-intent, higher bids)
cname = f'SP_Exact_Avento_Growth_{TIMESTAMP}'
agname = 'Avento_Exact_AG'
avento_skus = ['Avento-Black', 'Avento-Red', 'AVENTO-BLK-Blue', 'Avento-White', 'Avento-Blue', 'AVENTO-BLK-Red']
rows.append(make_campaign_row(cname, 40, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.65))
for sku in avento_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['polarized sports sunglasses', 'running sunglasses', 'cycling sunglasses men',
           'sport sunglasses uv400', 'polarized sunglasses for men', 'sports sunglasses for running']:
    rows.append(make_keyword_row(cname, agname, kw, 'exact', 0.65))
print(f'  Avento Exact: £40/day, 6 KW exact, 6 SKUs')

# 2. Avento ASIN Conquest
cname2 = f'SP_ASIN_Avento_Conquest_{TIMESTAMP}'
agname2 = 'Avento_ASIN_AG'
rows.append(make_campaign_row(cname2, 25, 'Manual'))
rows.append(make_ad_group_row(cname2, agname2, 0.55))
for sku in avento_skus:
    rows.append(make_product_ad_row(cname2, agname2, sku))
for target_asin in ['B07G8WDT6H', 'B07G8SL62W', 'B0C2P48NCR', 'B07G8W7FR8']:
    pt_row = {h: '' for h in SP_HEADERS}
    pt_row['Product'] = 'Sponsored Products'
    pt_row['Entity'] = 'Product Targeting'
    pt_row['Operation'] = 'Create'
    pt_row['Campaign ID'] = cname2
    pt_row['Ad Group ID'] = agname2
    pt_row['State'] = 'enabled'
    pt_row['Bid'] = 0.55
    pt_row['Product Targeting Expression'] = f'asin="{target_asin}"'
    rows.append(pt_row)
print(f'  Avento ASIN Conquest: £25/day, 4 PT targets')

# 3. Verano Growth
verano_skus = ['9L-KQGW-DL1C', 'QT-XAFB-5WKB', '4L-T4SW-6S0T', 'Y8-P4ZL-95RW', 'DT-7LPG-FVTB', 'F5-YYN7-BXR6']
cname = f'SP_Manual_Verano_Growth_{TIMESTAMP}'
agname = 'Verano_Growth_AG'
rows.append(make_campaign_row(cname, 20, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.50))
for sku in verano_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['polarized sports sunglasses', 'anti slip sunglasses', 'sunglasses men sports',
           'lightweight sunglasses men', 'running sunglasses uv400']:
    rows.append(make_keyword_row(cname, agname, kw, 'broad', 0.50))
for kw in ['polarized sports sunglasses', 'anti slip sunglasses']:
    rows.append(make_keyword_row(cname, agname, kw, 'exact', 0.55))
print(f'  Verano Growth: £20/day, 5 broad + 2 exact KW')

# 4. Dynamic Growth
dynamic_skus = ['0Q-W0KH-ONYA', 'M1-25EO-C8P7', 'ZH-EET1-OGHQ', 'EC-AYZ3-HZH4', '0P-KB7N-PEZ8']
cname = f'SP_Manual_Dynamic_Growth_{TIMESTAMP}'
agname = 'Dynamic_Growth_AG'
rows.append(make_campaign_row(cname, 15, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.45))
for sku in dynamic_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['polarized sports sunglasses', 'sunglasses men uv400', 'sports sunglasses cycling',
           'lightweight sport sunglasses']:
    rows.append(make_keyword_row(cname, agname, kw, 'broad', 0.45))
rows.append(make_keyword_row(cname, agname, 'polarized sports sunglasses', 'exact', 0.50))
print(f'  Dynamic Growth: £15/day, 4 broad + 1 exact KW')

# 5. Ventura Launch
ventura_skus = ['VT-BLK-R-Red 5', 'VT-BLK-R-Grey 5', 'VT-BLK-Y-Red 5', 'VT-BLK-Y-Blue 5', 'VT-WHT-Blue 5', 'VT-WHT-Grey 5']
cname = f'SP_Auto_Ventura_Launch_{TIMESTAMP}'
agname = 'Ventura_Auto_AG'
rows.append(make_campaign_row(cname, 10, 'Auto'))
rows.append(make_ad_group_row(cname, agname, 0.45))
for sku in ventura_skus:
    rows.append(make_product_ad_row(cname, agname, sku))

cname = f'SP_Manual_Ventura_Launch_{TIMESTAMP}'
agname = 'Ventura_Manual_AG'
rows.append(make_campaign_row(cname, 10, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.45))
for sku in ventura_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['sports sunglasses anti fog', 'performance sunglasses men', 'high performance sunglasses',
           'sunglasses cycling anti fog']:
    rows.append(make_keyword_row(cname, agname, kw, 'broad', 0.45))
print(f'  Ventura Launch: £20/day (Auto + Manual), 4 broad KW')

# 6. Laso-17 Launch
laso_skus = ['Laso-17 BLK-Red', 'Laso-17 BLK-Brown', 'Laso-17 BLK-Gold', 'Laso-17 BLK-Grey', 'Laso-17 BLK-Blue']
cname = f'SP_Auto_Laso17_Launch_{TIMESTAMP}'
agname = 'Laso17_Auto_AG'
rows.append(make_campaign_row(cname, 8, 'Auto'))
rows.append(make_ad_group_row(cname, agname, 0.40))
for sku in laso_skus:
    rows.append(make_product_ad_row(cname, agname, sku))

cname = f'SP_Manual_Laso17_Launch_{TIMESTAMP}'
agname = 'Laso17_Manual_AG'
rows.append(make_campaign_row(cname, 8, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.40))
for sku in laso_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['polarized sunglasses men', 'square sunglasses men', 'lightweight sunglasses']:
    rows.append(make_keyword_row(cname, agname, kw, 'broad', 0.40))
print(f'  Laso-17 Launch: £16/day (Auto + Manual), 3 broad KW')

# 7. Wire Strap Launch
wire_skus = ['Strap102-BLK/Grey', 'Strap102-BLK/BLK', 'Strap102-RED/BLK', 'Strap102-OG/BLK']
cname = f'SP_Auto_WireStrap_Launch_{TIMESTAMP}'
agname = 'WireStrap_Auto_AG'
rows.append(make_campaign_row(cname, 8, 'Auto'))
rows.append(make_ad_group_row(cname, agname, 0.33))
for sku in wire_skus:
    rows.append(make_product_ad_row(cname, agname, sku))

cname = f'SP_Manual_WireStrap_Launch_{TIMESTAMP}'
agname = 'WireStrap_Manual_AG'
rows.append(make_campaign_row(cname, 8, 'Manual'))
rows.append(make_ad_group_row(cname, agname, 0.33))
for sku in wire_skus:
    rows.append(make_product_ad_row(cname, agname, sku))
for kw in ['eyeglass strap', 'glasses strap', 'sunglass strap', 'glasses holder', 'eyeglass retainer']:
    rows.append(make_keyword_row(cname, agname, kw, 'broad', 0.33))
print(f'  Wire Strap Launch: £16/day (Auto + Manual), 5 broad KW')

# ═══════════════════════════════════════════════════════════════
# WRITE
# ═══════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
write_sheet(wb, 'Sponsored Products Campaigns', SP_HEADERS, rows, SP_ID_COLS)
path = '/Users/koda/amazon-autopilot/clients/flux/bulk-output/UK_SP_GrowthRebalance_20260310.xlsx'
wb.save(path)

pause_count = sum(1 for r in rows if r['Operation'] == 'Update' and r.get('State') == 'paused')
create_count = sum(1 for r in rows if r['Operation'] == 'Create' and r['Entity'] == 'Campaign')

print(f'\n=== SUMMARY ===')
print(f'File: {path}')
print(f'Total rows: {len(rows)}')
print(f'Campaigns PAUSED: {pause_count}')
print(f'Campaigns CREATED: {create_count} (£152/day new budget)')
print(f'Net effect: Stop bleeding ~£67/day wasted → Reinvest £152/day into proven performers')
