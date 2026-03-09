#!/usr/bin/env python3
"""
Generate Spring Deal warm-up SP campaign bulk files for UK and DE.

UK Deal Products: Avento, Jet II, Verano, Dynamic
DE Deal Products: Avento, Sportech, Jet II

Per Amazon Ads Bulk Checklist:
- SP Create = 4 rows per campaign: Campaign → Ad Group → Product Ad (with SKU) → Targeting
- Campaign ID = Campaign Name (for Create)
- Ad Group ID = Ad Group Name (for Create)
- Operation = 'Create' (capital C)
- Start Date = blank (State=enabled)
- Child ASINs only, with SKU
- Bidding Strategy = 'Dynamic bids - down only'
- Bid ≤ Daily Budget, Budget ≥ $1
- Campaign name includes timestamp to avoid conflicts
"""

import openpyxl
from openpyxl.utils import get_column_letter

TIMESTAMP = '20260310_1200'

# ─── SP Headers (52 columns, matching Amazon bulk template) ───
SP_HEADERS = [
    'Product', 'Entity', 'Operation', 'Campaign ID', 'Ad Group ID',
    'Portfolio ID', 'Ad ID', 'Keyword ID', 'Product Targeting ID',
    'Campaign Name', 'Ad Group Name',
    'Campaign Name (Informational only)', 'Ad Group Name (Informational only)',
    'Portfolio Name (Informational only)', 'Start Date', 'End Date',
    'Targeting Type', 'State',
    'Campaign State (Informational only)', 'Ad Group State (Informational only)',
    'Daily Budget', 'SKU', 'ASIN (Informational only)',
    'Eligibility Status (Informational only)',
    'Reason for Ineligibility (Informational only)',
    'Ad Group Default Bid', 'Ad Group Default Bid (Informational only)', 'Bid',
    'Keyword Text', 'Native Language Keyword', 'Native Language Locale',
    'Match Type', 'Bidding Strategy', 'Placement', 'Percentage',
    'Product Targeting Expression',
    'Resolved Product Targeting Expression (Informational only)',
    'Audience ID', 'Shopper Cohort Percentage', 'Shopper Cohort Type',
    'Segment Name (Informational only)',
    'Impressions', 'Clicks', 'Click-through Rate', 'Spend', 'Sales',
    'Orders', 'Units', 'Conversion Rate', 'ACOS', 'CPC', 'ROAS',
]

SP_ID_COLS = [3, 4, 5, 6, 7, 8]


def set_id_format(ws, id_col_indices, max_row):
    for col_idx in id_col_indices:
        col_letter = get_column_letter(col_idx + 1)
        for row in range(1, max_row + 1):
            ws[f'{col_letter}{row}'].number_format = '@'


def write_sheet(wb, sheet_name, headers, rows, id_col_indices):
    ws = wb.active
    ws.title = sheet_name
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, '')
            ws.cell(row=row_idx, column=col_idx, value=val)
    set_id_format(ws, id_col_indices, len(rows) + 1)


def make_campaign_row(campaign_name, daily_budget, targeting_type='Auto'):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Campaign'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Campaign Name'] = campaign_name
    row['State'] = 'enabled'
    row['Daily Budget'] = daily_budget
    row['Targeting Type'] = targeting_type
    row['Bidding Strategy'] = 'Dynamic bids - down only'
    # Start Date blank for SP with State=enabled
    return row


def make_ad_group_row(campaign_name, ad_group_name, default_bid):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Ad Group'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['Ad Group Name'] = ad_group_name
    row['State'] = 'enabled'
    row['Ad Group Default Bid'] = default_bid
    return row


def make_product_ad_row(campaign_name, ad_group_name, sku):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Product Ad'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['SKU'] = sku
    row['State'] = 'enabled'
    return row


def make_keyword_row(campaign_name, ad_group_name, keyword, match_type, bid):
    row = {h: '' for h in SP_HEADERS}
    row['Product'] = 'Sponsored Products'
    row['Entity'] = 'Keyword'
    row['Operation'] = 'Create'
    row['Campaign ID'] = campaign_name
    row['Ad Group ID'] = ad_group_name
    row['Keyword Text'] = keyword
    row['Match Type'] = match_type
    row['Bid'] = bid
    row['State'] = 'enabled'
    return row


# ═════════════════════════════════════════════════════════════════
# UK SPRING DEAL CAMPAIGNS
# ═════════════════════════════════════════════════════════════════

uk_rows = []

# --- UK Product Data (Child ASINs + SKUs from Products.csv) ---
UK_PRODUCTS = {
    'Avento': {
        'children': [
            ('B07G8WDT6H', 'Avento-Black'),
            ('B07G8SL62W', 'Avento-Red'),
            ('B0C2P48NCR', 'AVENTO-BLK-Blue'),
            ('B07G8W7FR8', 'Avento-White'),
            ('B07G8WNCK3', 'Avento-Blue'),
            ('B0C42LHKQ8', 'AVENTO-BLK-Red'),
        ],
        'budget': 57,     # from deal week target
        'bid': 0.55,      # sunglasses bid
        'keywords': [
            'polarized sports sunglasses',
            'running sunglasses men',
            'cycling sunglasses',
            'sports sunglasses uv400',
            'polarized sunglasses men',
        ],
    },
    'JetII': {
        'children': [
            ('B07VMRK3KV', 'JETII-BLK-GREY'),
            ('B08X9L1KRH', 'JETII-BLK-Blue'),
            ('B0DXPTBYZD', 'JETII-RED-GREY'),
        ],
        'budget': 18,
        'bid': 0.45,
        'keywords': [
            'lightweight polarized sunglasses',
            'sports sunglasses men',
            'polarized sunglasses cycling',
            'uv protection sunglasses',
        ],
    },
    'Verano': {
        'children': [
            ('B07CNKVWQP', '9L-KQGW-DL1C'),
            ('B07CNV788S', 'Y8-P4ZL-95RW'),
            ('B07CNSYJKD', 'QT-XAFB-5WKB'),
            ('B07CNYLXCC', '4L-T4SW-6S0T'),
            ('B07CNV8S82', 'DT-7LPG-FVTB'),
            ('B07CP3HSYD', 'F5-YYN7-BXR6'),
        ],
        'budget': 12,
        'bid': 0.40,
        'keywords': [
            'polarized sports sunglasses',
            'anti slip sunglasses',
            'sports sunglasses men women',
        ],
    },
    'Dynamic': {
        'children': [
            ('B07CNNXJ4Y', '0Q-W0KH-ONYA'),
            ('B07CNSD47H', 'M1-25EO-C8P7'),
            ('B07CNZPNN9', 'ZH-EET1-OGHQ'),
            ('B07CP26XD7', 'EC-AYZ3-HZH4'),
            ('B07CP46YKD', '0P-KB7N-PEZ8'),
        ],
        'budget': 8,
        'bid': 0.40,
        'keywords': [],  # Auto campaign only — no ad history
    },
}

# -- Generate UK campaigns --
for product_name, cfg in UK_PRODUCTS.items():
    # Campaign 1: SP Auto (all products get an auto campaign)
    auto_name = f'SP_Auto_{product_name}_SpringDeal_{TIMESTAMP}'
    auto_ag = f'{product_name}_Auto_AG'
    uk_rows.append(make_campaign_row(auto_name, cfg['budget'], 'Auto'))
    uk_rows.append(make_ad_group_row(auto_name, auto_ag, cfg['bid']))
    for asin, sku in cfg['children']:
        uk_rows.append(make_product_ad_row(auto_name, auto_ag, sku))

    # Campaign 2: SP Manual (only if keywords defined)
    if cfg['keywords']:
        manual_name = f'SP_Manual_{product_name}_SpringDeal_{TIMESTAMP}'
        manual_ag = f'{product_name}_Manual_AG'
        uk_rows.append(make_campaign_row(manual_name, cfg['budget'], 'Manual'))
        uk_rows.append(make_ad_group_row(manual_name, manual_ag, cfg['bid']))
        for asin, sku in cfg['children']:
            uk_rows.append(make_product_ad_row(manual_name, manual_ag, sku))
        for kw in cfg['keywords']:
            uk_rows.append(make_keyword_row(manual_name, manual_ag, kw, 'broad', cfg['bid']))

# Write UK file
wb_uk = openpyxl.Workbook()
write_sheet(wb_uk, 'Sponsored Products Campaigns', SP_HEADERS, uk_rows, SP_ID_COLS)
uk_path = '/Users/koda/amazon-autopilot/clients/flux/bulk-output/UK_SP_SpringDeal_20260310.xlsx'
wb_uk.save(uk_path)
print(f'Created: {uk_path}')
print(f'  UK rows: {len(uk_rows)}')


# ═════════════════════════════════════════════════════════════════
# DE SPRING DEAL CAMPAIGNS
# ═════════════════════════════════════════════════════════════════

de_rows = []

DE_PRODUCTS = {
    'Avento': {
        'children': [
            ('B07G8WDT6H', 'Avento-Black'),
            ('B07G8SL62W', 'Avento-Red'),
            ('B0C2P48NCR', 'AVENTO-BLK-Blue'),
            ('B07G8W7FR8', 'Avento-White'),
            ('B07G8WNCK3', 'Avento-Blue'),
            ('B0C42LHKQ8', 'AVENTO-BLK-Red'),
        ],
        'budget': 176,    # from deal week target (DE Avento is much bigger)
        'bid': 0.55,
        'keywords': [
            'polarisierte sportsonnenbrille',
            'sonnenbrille herren sport',
            'laufen sonnenbrille',
            'fahrrad sonnenbrille',
            'polarized sports sunglasses',
            'sonnenbrille uv400',
        ],
    },
    'Sportech': {
        'children': [
            ('B07CNN4FW7', 'Sportech-S.BLK-Grey'),
            ('B07CNNHJV5', 'Sportech-S.BLK-Red'),
            ('B07CNNMCBD', 'Sportech-M.BLK-Grey'),
            ('B07CNT173P', 'Sportech-S.BLK-Blue'),
        ],
        'budget': 22,
        'bid': 0.45,
        'keywords': [
            'sport sonnenbrille herren',
            'polarisierte sonnenbrille',
            'sonnenbrille sport damen herren',
            'laufbrille herren',
        ],
    },
    'JetII': {
        'children': [
            ('B07VMRK3KV', 'JETII-BLK-GREY'),
            ('B08X9L1KRH', 'JETII-BLK-Blue'),
            ('B0DXPTBYZD', 'JETII-RED-GREY'),
        ],
        'budget': 10,
        'bid': 0.40,
        'keywords': [],  # Auto only — zero W5 sales, need discovery
    },
}

for product_name, cfg in DE_PRODUCTS.items():
    # SP Auto
    auto_name = f'SP_Auto_{product_name}_SpringDeal_{TIMESTAMP}'
    auto_ag = f'{product_name}_Auto_AG'
    de_rows.append(make_campaign_row(auto_name, cfg['budget'], 'Auto'))
    de_rows.append(make_ad_group_row(auto_name, auto_ag, cfg['bid']))
    for asin, sku in cfg['children']:
        de_rows.append(make_product_ad_row(auto_name, auto_ag, sku))

    # SP Manual (only if keywords)
    if cfg['keywords']:
        manual_name = f'SP_Manual_{product_name}_SpringDeal_{TIMESTAMP}'
        manual_ag = f'{product_name}_Manual_AG'
        de_rows.append(make_campaign_row(manual_name, cfg['budget'], 'Manual'))
        de_rows.append(make_ad_group_row(manual_name, manual_ag, cfg['bid']))
        for asin, sku in cfg['children']:
            de_rows.append(make_product_ad_row(manual_name, manual_ag, sku))
        for kw in cfg['keywords']:
            de_rows.append(make_keyword_row(manual_name, manual_ag, kw, 'broad', cfg['bid']))

# Write DE file
wb_de = openpyxl.Workbook()
write_sheet(wb_de, 'Sponsored Products Campaigns', SP_HEADERS, de_rows, SP_ID_COLS)
de_path = '/Users/koda/amazon-autopilot/clients/flux/bulk-output/DE_SP_SpringDeal_20260310.xlsx'
wb_de.save(de_path)
print(f'Created: {de_path}')
print(f'  DE rows: {len(de_rows)}')

print('\n=== Summary ===')
print('UK: 4 products × (Auto + Manual where applicable)')
print('  Avento: Auto £57/day + Manual (5 KW broad)')
print('  Jet II: Auto £18/day + Manual (4 KW broad)')
print('  Verano: Auto £12/day + Manual (3 KW broad)')
print('  Dynamic: Auto £8/day only (no KW history)')
print('DE: 3 products × (Auto + Manual where applicable)')
print('  Avento: Auto €176/day + Manual (6 KW broad)')
print('  Sportech: Auto €22/day + Manual (4 KW broad)')
print('  Jet II: Auto €10/day only (no sales, discovery mode)')
print('\nBidding: Dynamic bids - down only | Sunglasses default: £/€0.55 | Lower tier: £/€0.40-0.45')
