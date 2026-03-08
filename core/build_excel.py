#!/usr/bin/env python3
"""
DAIKEN US — Weekly Performance Excel Report Builder
Full transparency: all raw data that powers the HTML dashboard
"""
import json, re
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as num_fmt)
from openpyxl.utils import get_column_letter

# ── Load data from generated HTML ─────────────────────────────────────────
HTML_PATH = "/Users/koda/amazon-autopilot/output/daiken/DAIKEN_US_Feb-Mar_2026.html"
OUTPUT_PATH = "/Users/koda/amazon-autopilot/output/daiken/20260307_DAIKEN_US_Feb-Mar_2026.xlsx"

with open(HTML_PATH, encoding="utf-8") as f:
    html = f.read()

parents_raw  = json.loads(re.search(r"const PARENTS = (\{.+?\});", html, re.DOTALL).group(1))
child_raw    = json.loads(re.search(r"const CHILD_DATA = (\[.+?\]);", html, re.DOTALL).group(1))
zh_raw       = json.loads(re.search(r"const ZH = (\{.+?\});", html).group(1))
zh_child_raw = json.loads(re.search(r"const CHILD_ZH = (\{.+?\});", html).group(1))
weeks        = json.loads(re.search(r"const WEEKS = (\[.+?\]);", html).group(1))
w_labels     = json.loads(re.search(r"const W_LABELS = (\{.+?\});", html).group(1))
tacos_target = int(re.search(r"const TACOS_TARGET = (\d+)", html).group(1))

SORTED_PARENTS = sorted(parents_raw.items(), key=lambda x: -x[1]["total_sales"])
NW = len(weeks)

# ── Style helpers ──────────────────────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color.lstrip("#"), end_color=hex_color.lstrip("#"))

def thin_border(top=False, bottom=False, left=False, right=False):
    t = Side(style="thin") if True else None
    return Border(
        top=Side(style="thin") if top else None,
        bottom=Side(style="thin") if bottom else None,
        left=Side(style="thin") if left else None,
        right=Side(style="thin") if right else None,
    )

HEADER_FILL  = hex_fill("#1e40af")   # Deep blue
WEEK_FILLS   = {
    "W1": hex_fill("#eff6ff"), "W2": hex_fill("#f0fdf4"),
    "W3": hex_fill("#fefce8"), "W4": hex_fill("#fff7ed"),
    "W5": hex_fill("#fdf4ff"),
}
WEEK_HEADER_FILLS = {
    "W1": hex_fill("#bfdbfe"), "W2": hex_fill("#bbf7d0"),
    "W3": hex_fill("#fef08a"), "W4": hex_fill("#fed7aa"),
    "W5": hex_fill("#e9d5ff"),
}
SUBHEADER_FILL = hex_fill("#f1f5f9")
WHITE_FILL     = hex_fill("#ffffff")
LIGHT_GREY     = hex_fill("#f8fafc")

BOLD_WHITE   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_DARK    = Font(name="Arial", bold=True, color="0f172a", size=10)
BOLD_SM      = Font(name="Arial", bold=True, color="0f172a", size=9)
NORMAL       = Font(name="Arial", size=9, color="0f172a")
MUTED        = Font(name="Arial", size=9, color="64748b")
LINK_FONT    = Font(name="Arial", size=9, color="2563eb", underline="single")
GREEN_BOLD   = Font(name="Arial", bold=True, size=9, color="15803d")
RED_BOLD     = Font(name="Arial", bold=True, size=9, color="dc2626")
AMBER_BOLD   = Font(name="Arial", bold=True, size=9, color="b45309")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

USD   = '#,##0.00'
USD0  = '#,##0'
PCT   = '0.0%'
INT   = '#,##0'
NEG   = '#,##0;(#,##0);-'
DASH_FMT = '#,##0;(#,##0);"-"'

def tacos_font(val):
    if val is None: return MUTED
    if val <= tacos_target: return GREEN_BOLD
    if val <= tacos_target * 1.2: return AMBER_BOLD
    return RED_BOLD

def tacos_fill(val, base_fill):
    """Keep base fill but we'll just use font color."""
    return base_fill

def set_cell(ws, row, col, value, font=None, fill=None, align=None,
             number_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if number_format: c.number_format = number_format
    if border: c.border = border
    return c

def merge_header(ws, row, col_start, col_end, text, fill, font=None, height=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill
    c.font = font or BOLD_WHITE
    c.alignment = CENTER
    if height: ws.row_dimensions[row].height = height

wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: Summary Dashboard
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("📊 5W Summary")
ws1.freeze_panes = "B4"
ws1.sheet_view.showGridLines = False

# Title row
ws1.merge_cells("A1:Q1")
c = ws1["A1"]
c.value = f"DAIKEN US · Weekly Performance Report  |  Feb–Mar 2026  |  W1–W{NW}  |  TACOS Target ≤{tacos_target}%"
c.font = Font(name="Arial", bold=True, size=12, color="0f172a")
c.fill = hex_fill("#f1f5f9")
c.alignment = LEFT
ws1.row_dimensions[1].height = 26

ws1.row_dimensions[2].height = 6  # spacer

# Column headers row 3
COLS = ["Product (ZH)", "ASIN"] + [f"{w}\n{w_labels[w]}" for w in weeks] + ["5W Total", "Avg/Week"]
HDR_COL_GROUPS = [("A", "Product"), ("B", "Parent ASIN")] + \
    [(get_column_letter(3+i), f"{weeks[i]} {w_labels[weeks[i]]}") for i in range(NW)] + \
    [(get_column_letter(3+NW), "5W Total"), (get_column_letter(4+NW), "Avg/Week")]

# Sub-metrics columns per week: Sales, Spend, TACOS, Units
# Layout: col A=Product, col B=ASIN,
#         then for each week: Sales(+1), Spend(+2), TACOS(+3), Units(+4)
#         then Total Sales, Total Spend, Total TACOS, Total Units
# Total cols: 2 + NW*4 + 4

COL_A = 1  # Product
COL_B = 2  # ASIN
def week_col(wi, metric):
    # wi = 0-indexed week index, metric 0=Sales 1=Spend 2=TACOS 3=Units
    return 3 + wi*4 + metric
def total_col(metric):
    return 3 + NW*4 + metric  # 4 total cols

TOTAL_COLS = total_col(0)  # first total col

# Row 3: week group headers (merged)
r3 = 3
ws1.row_dimensions[r3].height = 16
set_cell(ws1, r3, COL_A, "Product", BOLD_SM, SUBHEADER_FILL, CENTER)
set_cell(ws1, r3, COL_B, "Parent ASIN", BOLD_SM, SUBHEADER_FILL, CENTER)
for wi, w in enumerate(weeks):
    c_start = week_col(wi, 0)
    c_end   = week_col(wi, 3)
    ws1.merge_cells(start_row=r3, start_column=c_start, end_row=r3, end_column=c_end)
    cc = ws1.cell(row=r3, column=c_start)
    cc.value = f"{w}  ·  {w_labels[w]}"
    cc.font  = Font(name="Arial", bold=True, size=9, color="0f172a")
    cc.fill  = WEEK_HEADER_FILLS[w]
    cc.alignment = CENTER
ws1.merge_cells(start_row=r3, start_column=TOTAL_COLS, end_row=r3, end_column=TOTAL_COLS+3)
cc = ws1.cell(row=r3, column=TOTAL_COLS)
cc.value = "5W TOTAL"
cc.font = BOLD_SM
cc.fill = hex_fill("#1e3a5f")
cc.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
cc.alignment = CENTER

# Row 4: metric headers
r4 = 4
ws1.row_dimensions[r4].height = 20
set_cell(ws1, r4, COL_A, "產品名稱", BOLD_SM, SUBHEADER_FILL, LEFT)
set_cell(ws1, r4, COL_B, "Parent ASIN", BOLD_SM, SUBHEADER_FILL, CENTER)
metric_labels = ["Sales ($)", "Spend ($)", "TACOS", "Units"]
for wi in range(NW):
    fill = WEEK_HEADER_FILLS[weeks[wi]]
    for mi, lbl in enumerate(metric_labels):
        set_cell(ws1, r4, week_col(wi, mi), lbl, BOLD_SM, fill, CENTER)
total_metric_labels = ["Sales ($)", "Spend ($)", "TACOS", "Units"]
for mi, lbl in enumerate(total_metric_labels):
    set_cell(ws1, r4, TOTAL_COLS+mi, lbl, Font(name="Arial", bold=True, size=9, color="FFFFFF"),
             hex_fill("#1e3a5f"), CENTER)

# Data rows (one per parent)
data_start_row = 5
for ri, (pid, p) in enumerate(SORTED_PARENTS):
    row = data_start_row + ri
    ws1.row_dimensions[row].height = 18
    fill = LIGHT_GREY if ri % 2 == 0 else WHITE_FILL
    zh_name = zh_raw.get(pid, p["short"])
    set_cell(ws1, row, COL_A, zh_name, BOLD_DARK, fill, LEFT)
    set_cell(ws1, row, COL_B, pid, MUTED, fill, CENTER)

    for wi, w in enumerate(weeks):
        wd = p["weeks"].get(w, {})
        sales = wd.get("sales", 0) or 0
        spend = wd.get("spend", 0) or 0
        tacos = wd.get("tacos")
        units = wd.get("units", 0) or 0
        wfill = WEEK_FILLS[w]
        set_cell(ws1, row, week_col(wi,0), sales, NORMAL, wfill, RIGHT, USD0)
        set_cell(ws1, row, week_col(wi,1), spend, NORMAL, wfill, RIGHT, USD0)
        tc = ws1.cell(row=row, column=week_col(wi,2),
                      value=(tacos/100 if tacos is not None else None))
        tc.font = tacos_font(tacos); tc.fill = wfill; tc.alignment = RIGHT
        tc.number_format = PCT if tacos is not None else '@'
        set_cell(ws1, row, week_col(wi,3), units, NORMAL, wfill, RIGHT, INT)

    # Totals
    tot_s = p["total_sales"]; tot_e = p["total_spend"]
    tot_units = sum(p["weeks"].get(w,{}).get("units",0) or 0 for w in weeks)
    tot_tacos = (tot_e/tot_s) if tot_s > 0 else None
    tfill = hex_fill("#e0f2fe")
    set_cell(ws1, row, TOTAL_COLS+0, tot_s, BOLD_DARK, tfill, RIGHT, USD0)
    set_cell(ws1, row, TOTAL_COLS+1, tot_e, BOLD_DARK, tfill, RIGHT, USD0)
    tc = ws1.cell(row=row, column=TOTAL_COLS+2,
                  value=(tot_tacos/100 if tot_tacos else None))
    tc.font = tacos_font(tot_tacos); tc.fill = tfill; tc.alignment = RIGHT
    tc.number_format = PCT if tot_tacos else '@'
    set_cell(ws1, row, TOTAL_COLS+3, tot_units, BOLD_DARK, tfill, RIGHT, INT)

# Totals footer row
footer_row = data_start_row + len(SORTED_PARENTS)
ws1.row_dimensions[footer_row].height = 20
ff = hex_fill("#1e293b")
ws1.merge_cells(start_row=footer_row, start_column=COL_A, end_row=footer_row, end_column=COL_B)
set_cell(ws1, footer_row, COL_A, "ACCOUNT TOTAL", Font(name="Arial", bold=True, size=9, color="FFFFFF"), ff, LEFT)
for wi, w in enumerate(weeks):
    wd_sum_s  = sum((p["weeks"].get(w,{}).get("sales", 0) or 0) for _,p in SORTED_PARENTS)
    wd_sum_e  = sum((p["weeks"].get(w,{}).get("spend", 0) or 0) for _,p in SORTED_PARENTS)
    wd_sum_u  = sum((p["weeks"].get(w,{}).get("units", 0) or 0) for _,p in SORTED_PARENTS)
    wd_tacos  = (wd_sum_e/wd_sum_s) if wd_sum_s > 0 else None
    set_cell(ws1, footer_row, week_col(wi,0), wd_sum_s, Font(name="Arial",bold=True,size=9,color="FFFFFF"), ff, RIGHT, USD0)
    set_cell(ws1, footer_row, week_col(wi,1), wd_sum_e, Font(name="Arial",bold=True,size=9,color="FFFFFF"), ff, RIGHT, USD0)
    tc = ws1.cell(row=footer_row, column=week_col(wi,2), value=(wd_tacos/100 if wd_tacos else None))
    tc.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); tc.fill=ff; tc.alignment=RIGHT; tc.number_format=PCT if wd_tacos else '@'
    set_cell(ws1, footer_row, week_col(wi,3), wd_sum_u, Font(name="Arial",bold=True,size=9,color="FFFFFF"), ff, RIGHT, INT)
grand_s = sum(p["total_sales"] for _,p in SORTED_PARENTS)
grand_e = sum(p["total_spend"] for _,p in SORTED_PARENTS)
grand_u = sum(sum(p["weeks"].get(w,{}).get("units",0) or 0 for w in weeks) for _,p in SORTED_PARENTS)
grand_t = grand_e/grand_s if grand_s > 0 else None
for mi, val in enumerate([grand_s, grand_e, grand_t/100 if grand_t else None, grand_u]):
    cc = ws1.cell(row=footer_row, column=TOTAL_COLS+mi, value=val)
    cc.font=Font(name="Arial",bold=True,size=9,color="FFFFFF"); cc.fill=ff; cc.alignment=RIGHT
    cc.number_format = [USD0, USD0, PCT if grand_t else '@', INT][mi]

# Column widths
ws1.column_dimensions[get_column_letter(COL_A)].width = 18
ws1.column_dimensions[get_column_letter(COL_B)].width = 14
for wi in range(NW):
    for mi, w in enumerate([10, 9, 8, 7]):
        ws1.column_dimensions[get_column_letter(week_col(wi,mi))].width = w
for mi, w in enumerate([11, 10, 8, 7]):
    ws1.column_dimensions[get_column_letter(TOTAL_COLS+mi)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: Weekly Raw Data (all metrics, long format)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📋 Weekly Raw Data")
ws2.freeze_panes = "D3"
ws2.sheet_view.showGridLines = False

# Header
ws2.merge_cells("A1:P1")
c = ws2["A1"]
c.value = "Weekly Raw Data — All Metrics per Product per Week  (Source: ScaleInsight + Amazon Ads Console)"
c.font = Font(name="Arial", bold=True, size=11, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws2.row_dimensions[1].height = 24
ws2.row_dimensions[2].height = 4

# Column headers
r = 3
ws2.row_dimensions[r].height = 20
headers = [
    ("Product (ZH)",    12), ("Parent ASIN",  14), ("Week",  6), ("Period",   12),
    ("Sales ($)",       11), ("SP+SD Spend ($)", 13), ("SB Spend ($)", 12), ("Total Spend ($)", 13),
    ("Units",            7), ("Organic Units",  11), ("PPC Units",     9), ("Profits ($)",     11),
    ("TACOS (%)",        9), ("ACOS (approx)", 10), ("SP+SD ACOS",    10), ("SB Attr Sales",  12),
]
for ci, (hdr, width) in enumerate(headers, 1):
    set_cell(ws2, r, ci, hdr, BOLD_WHITE, HEADER_FILL, CENTER)
    ws2.column_dimensions[get_column_letter(ci)].width = width

# Data rows
row = 4
for pid, p in SORTED_PARENTS:
    zh_name = zh_raw.get(pid, p["short"])
    sb_attr_by_week = {}  # SB attr from total (approx: all in W5 for now)
    total_sb_attr = p.get("sb_attr", 0) or 0
    # Distribute sb_attr proportionally by sales per week
    total_sales_sum = p["total_sales"] or 1
    for w in weeks:
        w_sales = p["weeks"].get(w, {}).get("sales", 0) or 0
        sb_attr_by_week[w] = round(total_sb_attr * w_sales / total_sales_sum, 2)

    for w in weeks:
        wd = p["weeks"].get(w, {})
        sales   = wd.get("sales",   0) or 0
        spsd    = wd.get("spsd",    0) or 0
        sbsp    = wd.get("sbsp",    0) or 0
        spend   = wd.get("spend",   0) or 0
        units   = wd.get("units",   0) or 0
        organic = wd.get("organic", 0) or 0
        ppc     = wd.get("ppc",     0) or 0
        profits = wd.get("profits", 0) or 0
        tacos   = wd.get("tacos")
        acos    = (spend / (sales * (spend / spend)) * 100) if sales > 0 and spend > 0 else None
        spsd_acos = (spsd / sales * 100) if sales > 0 and spsd > 0 else None
        sb_attr_w = sb_attr_by_week.get(w, 0)

        fill = WEEK_FILLS[w]
        ws2.row_dimensions[row].height = 16
        set_cell(ws2, row, 1,  zh_name,  NORMAL, fill, LEFT)
        set_cell(ws2, row, 2,  pid,      MUTED,  fill, CENTER)
        set_cell(ws2, row, 3,  w,        BOLD_SM, fill, CENTER)
        set_cell(ws2, row, 4,  w_labels[w], MUTED, fill, CENTER)
        set_cell(ws2, row, 5,  sales,   NORMAL, fill, RIGHT, USD)
        set_cell(ws2, row, 6,  spsd,    NORMAL, fill, RIGHT, USD)
        set_cell(ws2, row, 7,  sbsp,    NORMAL, fill, RIGHT, USD)
        set_cell(ws2, row, 8,  spend,   NORMAL, fill, RIGHT, USD)
        set_cell(ws2, row, 9,  units,   NORMAL, fill, RIGHT, INT)
        set_cell(ws2, row, 10, organic, NORMAL, fill, RIGHT, INT)
        set_cell(ws2, row, 11, ppc,     NORMAL, fill, RIGHT, INT)
        set_cell(ws2, row, 12, profits, NORMAL if profits >= 0 else RED_BOLD, fill, RIGHT, USD)
        # TACOS
        tc = ws2.cell(row=row, column=13, value=tacos)
        tc.font=tacos_font(tacos); tc.fill=fill; tc.alignment=RIGHT
        tc.number_format = '0.0' if tacos else '@'
        set_cell(ws2, row, 14, acos,     MUTED,  fill, RIGHT, '0.0' if acos else '@')
        set_cell(ws2, row, 15, spsd_acos, MUTED, fill, RIGHT, '0.0' if spsd_acos else '@')
        set_cell(ws2, row, 16, sb_attr_w, MUTED, fill, RIGHT, USD)
        row += 1

    # Subtotal row per product
    ws2.row_dimensions[row].height = 18
    sf = hex_fill("#e2e8f0")
    set_cell(ws2, row, 1, f"  ↳ {zh_name} Total", BOLD_SM, sf, LEFT)
    set_cell(ws2, row, 2, pid, MUTED, sf, CENTER)
    set_cell(ws2, row, 3, f"W1–W{NW}", BOLD_SM, sf, CENTER)
    set_cell(ws2, row, 4, "Total", MUTED, sf, CENTER)
    # Use Excel SUM formulas pointing back to previous rows
    first_data_row = row - NW
    for ci, col in enumerate([5,6,7,8,9,10,11,12], 0):
        col_letter = get_column_letter(col)
        ws2.cell(row=row, column=col).value = f"=SUM({col_letter}{first_data_row}:{col_letter}{row-1})"
        ws2.cell(row=row, column=col).font = BOLD_DARK
        ws2.cell(row=row, column=col).fill = sf
        ws2.cell(row=row, column=col).alignment = RIGHT
        ws2.cell(row=row, column=col).number_format = USD if col in [5,6,7,8,12] else INT
    # TACOS total
    tc = ws2.cell(row=row, column=13)
    tc.value = f"=IF(E{row}>0,H{row}/E{row}*100,\"—\")"
    tc.font = BOLD_SM; tc.fill = sf; tc.alignment = RIGHT; tc.number_format = '0.0'
    set_cell(ws2, row, 16, total_sb_attr, BOLD_SM, sf, RIGHT, USD)
    row += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: Child ASIN Detail
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🔍 Child ASIN Detail")
ws3.freeze_panes = "D3"
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:M1")
c = ws3["A1"]
c.value = "Child ASIN Detail — Weekly Sales & Spend Breakdown"
c.font = Font(name="Arial", bold=True, size=11, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws3.row_dimensions[1].height = 24; ws3.row_dimensions[2].height = 4

r = 3
ws3.row_dimensions[r].height = 20
child_headers = [
    ("Parent Product",12), ("Child ASIN", 14), ("SKU/Name",     26), ("Week",  6),
    ("Sales ($)",     11), ("Spend ($)",   9),  ("Units",         7), ("Organic", 8),
    ("PPC",            7), ("Org %",       8),  ("ACOS (approx)", 11), ("5W Sales",10), ("5W Units",8),
]
for ci, (hdr, width) in enumerate(child_headers, 1):
    set_cell(ws3, r, ci, hdr, BOLD_WHITE, HEADER_FILL, CENTER)
    ws3.column_dimensions[get_column_letter(ci)].width = width

row = 4
# Group children by parent (in same order as sorted parents)
parent_order = [pid for pid, _ in SORTED_PARENTS]
children_by_parent = {pid: [] for pid in parent_order}
for c_item in child_raw:
    if c_item["parent"] in children_by_parent:
        children_by_parent[c_item["parent"]].append(c_item)

for pid in parent_order:
    p_data = parents_raw[pid]
    zh_parent = zh_raw.get(pid, p_data["short"])
    children = sorted(children_by_parent[pid], key=lambda x: -x["total_sales"])

    for c_item in children:
        asin = c_item["asin"]
        zh_name = zh_child_raw.get(asin, c_item.get("name", asin)[:40])
        c_total_sales = c_item["total_sales"]
        c_total_units = sum(c_item["weeks"].get(w,{}).get("units",0) or 0 for w in weeks)

        for w in weeks:
            wd = c_item["weeks"].get(w, {})
            sales   = wd.get("sales",   0) or 0
            spend   = wd.get("spend",   0) or 0
            units   = wd.get("units",   0) or 0
            organic = wd.get("organic", 0) or 0
            ppc     = units - organic
            org_pct = (organic / units) if units > 0 else None
            acos    = (spend / sales * 100) if sales > 0 and spend > 0 else None

            fill = WEEK_FILLS[w]
            ws3.row_dimensions[row].height = 16
            set_cell(ws3, row, 1, zh_parent,    MUTED,   fill, LEFT)
            set_cell(ws3, row, 2, asin,          MUTED,   fill, CENTER)
            set_cell(ws3, row, 3, zh_name,       NORMAL,  fill, LEFT)
            set_cell(ws3, row, 4, w,             BOLD_SM, fill, CENTER)
            set_cell(ws3, row, 5, sales,         NORMAL,  fill, RIGHT, USD)
            set_cell(ws3, row, 6, spend,         NORMAL,  fill, RIGHT, USD)
            set_cell(ws3, row, 7, units,         NORMAL,  fill, RIGHT, INT)
            set_cell(ws3, row, 8, organic,       NORMAL,  fill, RIGHT, INT)
            set_cell(ws3, row, 9, ppc if ppc >= 0 else 0, NORMAL, fill, RIGHT, INT)
            op = ws3.cell(row=row, column=10, value=org_pct)
            op.font=NORMAL; op.fill=fill; op.alignment=RIGHT
            op.number_format = PCT if org_pct is not None else '@'
            ac = ws3.cell(row=row, column=11, value=acos)
            ac.font=MUTED; ac.fill=fill; ac.alignment=RIGHT
            ac.number_format = '0.0' if acos else '@'
            set_cell(ws3, row, 12, c_total_sales,  MUTED, fill, RIGHT, USD0)
            set_cell(ws3, row, 13, c_total_units,  MUTED, fill, RIGHT, INT)
            row += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: WoW Week-over-Week Change
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📈 WoW Change")
ws4.freeze_panes = "B4"
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:L1")
c = ws4["A1"]
c.value = f"Week-over-Week Change  |  {NW-1} Comparison Periods  |  Green = Improvement, Red = Decline"
c.font = Font(name="Arial", bold=True, size=11, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws4.row_dimensions[1].height = 24; ws4.row_dimensions[2].height = 4

# For each consecutive week pair
periods = [(weeks[i], weeks[i+1]) for i in range(len(weeks)-1)]
N_PERIODS = len(periods)

ws4.row_dimensions[3].height = 16; ws4.row_dimensions[4].height = 20

# Row 3: period group headers
r3 = 3
set_cell(ws4, r3, 1, "產品",   BOLD_SM, SUBHEADER_FILL, LEFT)
set_cell(ws4, r3, 2, "Metric", BOLD_SM, SUBHEADER_FILL, LEFT)
WOW_FILLS = [hex_fill("#dbeafe"), hex_fill("#dcfce7"), hex_fill("#fef9c3"), hex_fill("#ffe4e6")]
for pi, (w_from, w_to) in enumerate(periods):
    c_start = 3 + pi*3
    ws4.merge_cells(start_row=3, start_column=c_start, end_row=3, end_column=c_start+2)
    cc = ws4.cell(row=3, column=c_start)
    cc.value = f"{w_from} → {w_to}"
    cc.font = BOLD_SM; cc.fill = WOW_FILLS[pi % len(WOW_FILLS)]; cc.alignment = CENTER

# Row 4: sub-headers for each period: From, To, Δ%
r4 = 4
set_cell(ws4, r4, 1, "產品名稱", BOLD_SM, SUBHEADER_FILL, LEFT)
set_cell(ws4, r4, 2, "Metric",  BOLD_SM, SUBHEADER_FILL, LEFT)
ws4.column_dimensions["A"].width = 16; ws4.column_dimensions["B"].width = 10
for pi, (w_from, w_to) in enumerate(periods):
    c_start = 3 + pi*3
    fill = WOW_FILLS[pi % len(WOW_FILLS)]
    set_cell(ws4, r4, c_start,   w_labels[w_from], BOLD_SM, fill, RIGHT)
    set_cell(ws4, r4, c_start+1, w_labels[w_to],   BOLD_SM, fill, RIGHT)
    set_cell(ws4, r4, c_start+2, "Δ%",             BOLD_SM, fill, CENTER)
    for di in range(3):
        ws4.column_dimensions[get_column_letter(c_start+di)].width = 10

# Data
row = 5
metrics = [("Sales", "sales"), ("Spend", "spend"), ("Units", "units"), ("TACOS", "tacos")]
for pid, p in SORTED_PARENTS:
    zh_name = zh_raw.get(pid, p["short"])
    first_metric_row = row
    for mi, (m_label, m_key) in enumerate(metrics):
        fill = LIGHT_GREY if row % 2 == 0 else WHITE_FILL
        ws4.row_dimensions[row].height = 16
        if mi == 0:
            set_cell(ws4, row, 1, zh_name, BOLD_DARK, fill, LEFT)
        else:
            set_cell(ws4, row, 1, "", NORMAL, fill, LEFT)
        set_cell(ws4, row, 2, m_label, MUTED, fill, LEFT)

        for pi, (w_from, w_to) in enumerate(periods):
            c_start = 3 + pi*3
            pfill = WOW_FILLS[pi % len(WOW_FILLS)]
            v_from = p["weeks"].get(w_from, {}).get(m_key) or 0
            v_to   = p["weeks"].get(w_to,   {}).get(m_key) or 0
            delta_pct = ((v_to - v_from) / v_from) if v_from and v_from != 0 else None

            fmt_from = USD0 if m_key in ["sales","spend"] else (INT if m_key=="units" else '0.0')
            fmt_to   = fmt_from
            set_cell(ws4, row, c_start,   v_from, NORMAL, pfill, RIGHT, fmt_from)
            set_cell(ws4, row, c_start+1, v_to,   NORMAL, pfill, RIGHT, fmt_to)
            dp = ws4.cell(row=row, column=c_start+2, value=delta_pct)
            dp.number_format = PCT if delta_pct is not None else '@'
            dp.alignment = CENTER; dp.fill = pfill
            if delta_pct is not None:
                # Sales/Units up = good; Spend/TACOS down = good
                is_good = (delta_pct > 0) if m_key in ["sales","units"] else (delta_pct < 0)
                if abs(delta_pct) < 0.02:
                    dp.font = MUTED
                elif is_good:
                    dp.font = GREEN_BOLD
                else:
                    dp.font = RED_BOLD
            else:
                dp.font = MUTED
        row += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: SB Attribution Detail
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("🎯 SB Attribution")
ws5.freeze_panes = "B4"
ws5.sheet_view.showGridLines = False

# Re-read SB CSV — full 5W
import csv as _csv, re as re2
from datetime import datetime, date
from collections import defaultdict as _dd

SB_CSV = "/Users/koda/amazon-autopilot/clients/daiken/input/SB_report.csv"

def _parse_money(v):
    return float(re2.sub(r"[^\d.]", "", str(v).strip())) if v and str(v).strip() else 0.0

def _parse_sb_date(s):
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try: return datetime.strptime(s.strip().strip('"'), fmt).date()
        except: pass
    return None

# Build week lookup from HTML data
from datetime import date as _date
_WEEK_RANGES = {
    "W1": (_date(2026,2,1),  _date(2026,2,7)),
    "W2": (_date(2026,2,8),  _date(2026,2,14)),
    "W3": (_date(2026,2,15), _date(2026,2,21)),
    "W4": (_date(2026,2,22), _date(2026,2,28)),
    "W5": (_date(2026,3,1),  _date(2026,3,7)),
}
def _get_week(d):
    for w,(s,e) in _WEEK_RANGES.items():
        if s <= d <= e: return w
    return "—"

# Load Products for ASIN → parent mapping
child_to_parent_map = {}
asin_to_name_map = {}
with open("/Users/koda/amazon-autopilot/clients/daiken/input/Products.csv", encoding="utf-8-sig") as f:
    for r_row in _csv.DictReader(f):
        child_to_parent_map[r_row["ASIN"].strip()] = r_row["Parent ASIN"].strip()
        asin_to_name_map[r_row["ASIN"].strip()] = r_row["Title"].strip()[:50]

sb_rows = []
SALES_COL_NAME = None
with open(SB_CSV, encoding="utf-8-sig") as f:
    for r_row in _csv.DictReader(f):
        if SALES_COL_NAME is None:
            SALES_COL_NAME = next((k for k in r_row.keys() if "14 Day Total Sales" in k and "Click" not in k), "14 Day Total Sales ")
        d = _parse_sb_date(r_row.get("Date",""))
        if not d: continue
        asin   = r_row.get("Purchased ASIN","").strip()
        s_val  = _parse_money(r_row.get(SALES_COL_NAME, 0))
        orders = int(_parse_money(r_row.get("14 Day Total Orders (#)", 0)))
        units  = int(_parse_money(r_row.get("14 Day Total Units (#)", 0)))
        ntb_s  = _parse_money(r_row.get("14 Day New-to-brand Sales", 0))
        ntb_o  = int(_parse_money(r_row.get("14 Day New-to-brand Orders (#)", 0)))
        parent = child_to_parent_map.get(asin, asin)
        zh_p   = zh_raw.get(parent, parent)
        week   = _get_week(d)
        camp   = r_row.get("Campaign Name","").replace("_ERIC","")
        sb_rows.append((d, week, parent, zh_p, asin, camp, s_val, orders, units, ntb_s, ntb_o))

sb_rows.sort(key=lambda x: (x[0], x[2]))

# Title
ws5.merge_cells("A1:K1")
c = ws5["A1"]
total_sb_sales = sum(r[6] for r in sb_rows)
total_sb_ntb   = sum(r[9] for r in sb_rows)
c.value = (f"Sponsored Brands Attributed Purchases — W1–W{len(_WEEK_RANGES)}  |  "
           f"{len(sb_rows)} rows  |  Total SB Sales: ${total_sb_sales:,.0f}  |  NTB: ${total_sb_ntb:,.0f}")
c.font = Font(name="Arial", bold=True, size=11, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws5.row_dimensions[1].height = 24

# Per-week summary bar (row 2)
ws5.row_dimensions[2].height = 18
week_sb_totals = _dd(lambda: {"sales":0,"orders":0,"ntb":0})
for r_sb in sb_rows:
    week_sb_totals[r_sb[1]]["sales"]  += r_sb[6]
    week_sb_totals[r_sb[1]]["orders"] += r_sb[7]
    week_sb_totals[r_sb[1]]["ntb"]    += r_sb[9]
col_c = 1
for w in ["W1","W2","W3","W4","W5"]:
    wd = week_sb_totals[w]
    ws5.merge_cells(start_row=2, start_column=col_c, end_row=2, end_column=col_c+1)
    wc = ws5.cell(row=2, column=col_c)
    wc.value = f"{w}: ${wd['sales']:,.0f} ({wd['orders']} ord)"
    wc.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    wc.fill = list(WEEK_HEADER_FILLS.values())[list(_WEEK_RANGES.keys()).index(w)]
    wc.font = Font(name="Arial", bold=True, size=8, color="0f172a")
    wc.alignment = CENTER
    col_c += 2
ws5.row_dimensions[3].height = 4  # spacer

# Headers row 4
r = 4
ws5.row_dimensions[r].height = 20
sb_headers = [
    ("Week",6), ("Date",11), ("產品 ZH",14), ("Parent ASIN",14), ("Child ASIN",14),
    ("Campaign",38), ("SB Sales ($)",13), ("Orders",8), ("Units",7),
    ("NTB Sales ($)",13), ("NTB Orders",10),
]
for ci, (hdr, width) in enumerate(sb_headers, 1):
    set_cell(ws5, r, ci, hdr, BOLD_WHITE, HEADER_FILL, CENTER)
    ws5.column_dimensions[get_column_letter(ci)].width = width

# Data rows grouped by week
row = 5
for w in ["W1","W2","W3","W4","W5"]:
    w_rows = [x for x in sb_rows if x[1] == w]
    if not w_rows: continue
    wfill = WEEK_FILLS.get(w, WHITE_FILL)
    whfill = WEEK_HEADER_FILLS.get(w, SUBHEADER_FILL)
    # Week section header
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    wh = ws5.cell(row=row, column=1)
    wd = week_sb_totals[w]
    wh.value = (f"  {w} — {w_labels.get(w,'')}  |  {len(w_rows)} 筆  |  "
                f"Sales ${wd['sales']:,.0f}  Orders {wd['orders']}  NTB ${wd['ntb']:,.0f}")
    wh.font = Font(name="Arial", bold=True, size=9, color="0f172a")
    wh.fill = whfill; wh.alignment = LEFT
    ws5.row_dimensions[row].height = 16
    row += 1
    for r_sb in w_rows:
        d_sb, week_sb, parent, zh_p, asin, camp, s_val, orders, units, ntb_s, ntb_o = r_sb
        ws5.row_dimensions[row].height = 15
        fill2 = wfill
        ntb_pct = ntb_s / s_val if s_val > 0 else None
        set_cell(ws5, row, 1,  week_sb,              MUTED,   fill2, CENTER)
        set_cell(ws5, row, 2,  d_sb.strftime("%m/%d"), NORMAL, fill2, CENTER, "@")
        set_cell(ws5, row, 3,  zh_p,                 BOLD_SM, fill2, LEFT)
        set_cell(ws5, row, 4,  parent,               MUTED,   fill2, CENTER)
        set_cell(ws5, row, 5,  asin,                 MUTED,   fill2, CENTER)
        set_cell(ws5, row, 6,  camp,                 NORMAL,  fill2, LEFT)
        set_cell(ws5, row, 7,  s_val,                NORMAL,  fill2, RIGHT, USD)
        set_cell(ws5, row, 8,  orders,               NORMAL,  fill2, RIGHT, INT)
        set_cell(ws5, row, 9,  units,                NORMAL,  fill2, RIGHT, INT)
        set_cell(ws5, row, 10, ntb_s,                NORMAL,  fill2, RIGHT, USD)
        set_cell(ws5, row, 11, ntb_o,                MUTED,   fill2, RIGHT, INT)
        row += 1

# Grand total footer
ws5.row_dimensions[row].height = 20
sf2 = hex_fill("#1e293b")
ws5.merge_cells(f"A{row}:F{row}")
set_cell(ws5, row, 1, f"5W GRAND TOTAL  ({len(sb_rows)} rows)", Font(name="Arial",bold=True,size=9,color="FFFFFF"), sf2, LEFT)
set_cell(ws5, row, 7, total_sb_sales, Font(name="Arial",bold=True,size=9,color="FFFFFF"), sf2, RIGHT, USD)
set_cell(ws5, row, 8, sum(r[7] for r in sb_rows), Font(name="Arial",bold=True,size=9,color="FFFFFF"), sf2, RIGHT, INT)
set_cell(ws5, row, 9, sum(r[8] for r in sb_rows), Font(name="Arial",bold=True,size=9,color="FFFFFF"), sf2, RIGHT, INT)
set_cell(ws5, row, 10, total_sb_ntb, Font(name="Arial",bold=True,size=9,color="FFFFFF"), sf2, RIGHT, USD)
ntb_grand = ws5.cell(row=row, column=11)
ntb_grand.value = total_sb_ntb / total_sb_sales if total_sb_sales > 0 else None
ntb_grand.font = Font(name="Arial",bold=True,size=9,color="FFFFFF"); ntb_grand.fill = sf2
ntb_grand.alignment = RIGHT; ntb_grand.number_format = PCT

# ── Sheet tab colors ─────────────────────────────────────────────────────
ws1.sheet_properties.tabColor = "1e40af"
ws2.sheet_properties.tabColor = "0f172a"
ws3.sheet_properties.tabColor = "7c3aed"
ws4.sheet_properties.tabColor = "16a34a"
ws5.sheet_properties.tabColor = "dc2626"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6: 產品象限 (Quadrant Analysis)
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("🎯 產品象限")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = "16a34a"

# Compute quadrant logic (mirrors HTML buildQuadTab)
sp_sorted = sorted(parents_raw.items(), key=lambda x: -x[1]["total_sales"])
sp_active = [(pid,p) for pid,p in sp_sorted if p["total_sales"] > 0]
sales_arr = sorted(p["total_sales"] for _,p in sp_active)
n = len(sales_arr)
med = (sales_arr[n//2-1]+sales_arr[n//2])/2 if n%2==0 else sales_arr[n//2]

def get_quadrant(p):
    s, t = p["total_sales"], p.get("tacos_4w") or 999
    if s >= med and t <= tacos_target:  return "star",     "⭐ Star",     "16a34a", "高銷售 · 低 TACOS"
    if s >= med and t  > tacos_target:  return "question", "⚠️ Question", "d97706", "高銷售 · 高 TACOS"
    if s  < med and t <= tacos_target:  return "potential","💤 Potential","2563eb", "低銷售 · 低 TACOS"
    return                                     "cut",      "🔴 Cut",      "dc2626", "低銷售 · 高 TACOS"

QUAD_ACTIONS = {
    "star":     ["複製現有 SP 活動 Bulk File，新命名加 _COPY 後綴",
                 "全部 SP 活動預算 × 1.3（+30%）→ 上傳 Bulk File",
                 "確認是否有 SB Banner 廣告，若無則新建",
                 "建立 SD Retargeting 活動 targeting 已瀏覽 ASIN 受眾"],
    "question": ["下載 SP Bulk → ACOS ≥85% 關鍵字 Bid × 0.8（降 20%）",
                 "TACOS ≥85% 且 14 天無轉換關鍵字 → 改為 Paused",
                 "SP 活動預算維持不變，待 TACOS 下降再增",
                 "檢查競品定價，評估是否調整售價提升轉換率"],
    "cut":      ["Seller Central 售價下調 $0.50，觀察轉換率",
                 "建立 Coupon 折扣（5–10%）提升搜尋曝光",
                 "申請 Prime Exclusive Discount / Lightning Deal",
                 "SP Bulk File → 全部 Keyword Bid 降至 $2.00 上限"],
    "potential":["新建 SP Auto campaign，日預算 $10，CPC $2.00",
                 "新建 SB Banner campaign 指向 Store Page，CPC $2.00",
                 "新建 SD Audience campaign → Competitor ASIN 瀏覽受眾",
                 "建立 AMC Lookalike 受眾 → 套用至 SD campaign",
                 "用高轉換 KW 根詞建立 SP Exact Match campaign"],
}

QUAD_ORDER = ["star", "question", "cut", "potential"]
QUAD_COLORS = {"star":"e6f4ea","question":"fff8e1","cut":"fdecea","potential":"e3f2fd"}
QUAD_HEADER_COLORS = {"star":"16a34a","question":"d97706","cut":"dc2626","potential":"2563eb"}

# Title
ws6.merge_cells("A1:K1")
c = ws6["A1"]
c.value = f"產品象限分析  |  Sales Median = ${med:,.0f}  |  TACOS Target ≤{tacos_target}%  |  {NW}W Total"
c.font = Font(name="Arial", bold=True, size=12, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws6.row_dimensions[1].height = 26
ws6.row_dimensions[2].height = 6

# Summary KPI row
r = 3
quad_counts = {q: [] for q in QUAD_ORDER}
for pid, p in sp_active:
    q,_,_,_ = get_quadrant(p)
    quad_counts[q].append((pid, p))

kpi_labels = {"star":"⭐ Star","question":"⚠️ Question","cut":"🔴 Cut","potential":"💤 Potential"}
for qi, q in enumerate(QUAD_ORDER):
    col = 1 + qi*2
    ws6.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
    cc = ws6.cell(row=r, column=col)
    cc.value = f"{kpi_labels[q]}  ({len(quad_counts[q])} 個產品)"
    cc.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cc.fill = hex_fill(QUAD_HEADER_COLORS[q]); cc.alignment = CENTER
ws6.row_dimensions[r].height = 20
ws6.row_dimensions[4].height = 6

# Quadrant data table — headers
r = 5
ws6.row_dimensions[r].height = 18
hdr_cols = ["象限","產品名稱","Parent ASIN","5W Sales ($)","5W Spend ($)","5W TACOS","Sales Rank","vs Median","SB Attr Sales","建議動作方向","優先級"]
col_widths = [12, 16, 14, 12, 12, 9, 9, 11, 13, 30, 10]
for ci, (hdr, width) in enumerate(zip(hdr_cols, col_widths), 1):
    set_cell(ws6, r, ci, hdr, BOLD_WHITE, HEADER_FILL, CENTER)
    ws6.column_dimensions[get_column_letter(ci)].width = width

# Data rows grouped by quadrant
row = 6
for q in QUAD_ORDER:
    items = quad_counts[q]
    if not items: continue
    q_fill = hex_fill(QUAD_COLORS[q])
    q_hfill = hex_fill(QUAD_HEADER_COLORS[q])
    # Quadrant section header
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdr_cols))
    cc = ws6.cell(row=row, column=1)
    _,qlabel,_,qdesc = get_quadrant(items[0][1]) if items else ("","","","")
    cc.value = f"  {qlabel}  ·  {qdesc}  ({len(items)} 個產品)"
    cc.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cc.fill = q_hfill; cc.alignment = LEFT
    ws6.row_dimensions[row].height = 16
    row += 1

    for rank_i, (pid, p) in enumerate(sorted(items, key=lambda x: -x[1]["total_sales"]), 1):
        ws6.row_dimensions[row].height = 16
        zh = zh_raw.get(pid, p["short"])
        tot_s = p["total_sales"]; tot_e = p["total_spend"]
        tot_t = p.get("tacos_4w")
        sb_attr = p.get("sb_attr", 0) or 0
        vs_med = (tot_s - med) / med if med > 0 else 0
        priority = "本週執行" if q in ("star","question","cut") else "下週執行"
        action_summary = {"star":"擴大預算 · 複製 SP · 加 SB","question":"降 CPC · 暫停高 ACOS KW",
                          "cut":"降價 · Coupon · 降 Bid","potential":"新建 SP/SB/SD · AMC Lookalike"}[q]
        _,qlbl,_,_ = get_quadrant(p)
        set_cell(ws6, row, 1,  qlbl,      Font(name="Arial",bold=True,size=9,color=QUAD_HEADER_COLORS[q]), q_fill, CENTER)
        set_cell(ws6, row, 2,  zh,        BOLD_DARK, q_fill, LEFT)
        set_cell(ws6, row, 3,  pid,       MUTED,     q_fill, CENTER)
        set_cell(ws6, row, 4,  tot_s,     NORMAL,    q_fill, RIGHT, USD0)
        set_cell(ws6, row, 5,  tot_e,     NORMAL,    q_fill, RIGHT, USD0)
        tc = ws6.cell(row=row, column=6, value=(tot_t/100) if tot_t else None)
        tc.font=tacos_font(tot_t); tc.fill=q_fill; tc.alignment=RIGHT; tc.number_format=PCT if tot_t else '@'
        set_cell(ws6, row, 7,  rank_i,    MUTED,     q_fill, CENTER, INT)
        vm = ws6.cell(row=row, column=8, value=vs_med)
        vm.font = GREEN_BOLD if vs_med >= 0 else RED_BOLD; vm.fill=q_fill; vm.alignment=RIGHT; vm.number_format='+0.0%;-0.0%;0.0%'
        set_cell(ws6, row, 9,  sb_attr,   MUTED,     q_fill, RIGHT, USD0)
        set_cell(ws6, row, 10, action_summary, NORMAL, q_fill, LEFT)
        pf = ws6.cell(row=row, column=11, value=priority)
        pf.font = GREEN_BOLD if priority=="本週執行" else AMBER_BOLD; pf.fill=q_fill; pf.alignment=CENTER
        row += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7: 行動方案 (Action Plan)
# ════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("✅ 行動方案")
ws7.sheet_view.showGridLines = False
ws7.sheet_properties.tabColor = "7c3aed"

FULL_TASKS = {
    "star": [
        ("SP Bulk",  "本週", "下載現有 SP 活動 Bulk File → 複製所有 Enabled campaigns → 新命名加 _COPY 後綴"),
        ("Budget",   "本週", "全部 SP 活動預算 × 1.3（+30%）→ 上傳 Bulk File → 確認無錯誤"),
        ("SB",       "本週", "確認是否有 SB Banner 廣告，若無則新建 SB Campaign 指向此產品"),
        ("SD",       "本週", "建立 SD Retargeting 活動 targeting 已瀏覽 ASIN 的受眾"),
    ],
    "question": [
        ("CPC",      "本週", "下載 SP Bulk File → 找出 ACOS ≥85% 的關鍵字 → Bid × 0.8（降 20%）"),
        ("Pause KW", "本週", "TACOS ≥85% 且過去 14 天無轉換的關鍵字 → 狀態改為 Paused"),
        ("Budget",   "本週", "SP 活動預算維持不變，待 TACOS 下降再考慮增加"),
        ("Review",   "本週", "檢查競品 listing 定價，評估是否需要調整售價提高轉換率"),
    ],
    "cut": [
        ("Price",    "本週", "在 Seller Central 將售價下調 $0.50，觀察轉換率變化"),
        ("Coupon",   "本週", "建立 Coupon 折扣（建議 5–10%）→ 提升搜尋結果曝光"),
        ("Deal",     "本週", "申請 Prime Exclusive Discount 或 Lightning Deal（需提前 2 週申請）"),
        ("SP Bid",   "本週", "下載 SP Bulk File → 全部 Keyword Bid 降至 $2.00 上限"),
    ],
    "potential": [
        ("SP",       "下週", "新建 SP Auto campaign，日預算 $10，CPC 出價 $2.00，30天跑數據"),
        ("SB",       "下週", "新建 SB Banner campaign，指向此產品 Store Page，CPC $2.00"),
        ("SD",       "下週", "新建 SD Audience campaign，目標 Competitor ASIN 的瀏覽受眾"),
        ("AMC",      "下週", "建立 AMC Lookalike 受眾 → 相似購買行為 → 套用至 SD campaign"),
        ("KW",       "下週", "用現有暢銷產品的高轉換 KW 根詞，建立 SP Exact Match campaign"),
    ],
}

# Title
ws7.merge_cells("A1:H1")
c = ws7["A1"]
c.value = f"行動方案清單  |  DAIKEN US  |  W{NW} {w_labels[weeks[-1]]}  |  共 {sum(len(v) for v in quad_counts.values())} 個產品"
c.font = Font(name="Arial", bold=True, size=12, color="0f172a")
c.fill = hex_fill("#f1f5f9"); c.alignment = LEFT
ws7.row_dimensions[1].height = 26
ws7.row_dimensions[2].height = 6

r = 3
ws7.row_dimensions[r].height = 20
action_headers = [("完成","完成?",5),("產品",  "產品名稱",   16),("象限","象限",      12),
                  ("優先","優先級",   10),("類型","類型",       9),("任務","具體行動",  52),
                  ("銷售","5W Sales",12),("TACOS","TACOS",     9)]
for ci, (_, hdr, width) in enumerate(action_headers, 1):
    set_cell(ws7, r, ci, hdr, BOLD_WHITE, HEADER_FILL, CENTER)
    ws7.column_dimensions[get_column_letter(ci)].width = width

row = 4
for q in QUAD_ORDER:
    items = quad_counts[q]
    if not items: continue
    q_hfill = hex_fill(QUAD_HEADER_COLORS[q])
    # Section header
    ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cc = ws7.cell(row=row, column=1)
    _,qlabel,_,qdesc = get_quadrant(items[0][1])
    cc.value = f"  {qlabel}  —  {qdesc}"
    cc.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    cc.fill = q_hfill; cc.alignment = LEFT
    ws7.row_dimensions[row].height = 15
    row += 1

    q_fill = hex_fill(QUAD_COLORS[q])
    tasks = FULL_TASKS[q]
    priority_label = "本週執行" if q in ("star","question","cut") else "下週執行"
    for pid, p in sorted(items, key=lambda x: -x[1]["total_sales"]):
        zh = zh_raw.get(pid, p["short"])
        _,qlbl,_,_ = get_quadrant(p)
        for ti, (task_type, when, task_text) in enumerate(tasks):
            ws7.row_dimensions[row].height = 15
            # Checkbox column — leave blank, user ticks it
            set_cell(ws7, row, 1, "", NORMAL, q_fill, CENTER)
            if ti == 0:
                set_cell(ws7, row, 2, zh, BOLD_DARK, q_fill, LEFT)
                set_cell(ws7, row, 3, qlbl, Font(name="Arial",bold=True,size=9,color=QUAD_HEADER_COLORS[q]), q_fill, CENTER)
                pf2 = ws7.cell(row=row, column=4, value=priority_label)
                pf2.font = GREEN_BOLD if priority_label=="本週執行" else AMBER_BOLD
                pf2.fill=q_fill; pf2.alignment=CENTER
                tot_s2 = p["total_sales"]; tot_t2 = p.get("tacos_4w")
                set_cell(ws7, row, 7, tot_s2, NORMAL, q_fill, RIGHT, USD0)
                tc2 = ws7.cell(row=row, column=8, value=(tot_t2/100) if tot_t2 else None)
                tc2.font=tacos_font(tot_t2); tc2.fill=q_fill; tc2.alignment=RIGHT
                tc2.number_format=PCT if tot_t2 else '@'
            else:
                for ci in [2,3,4,7,8]:
                    ws7.cell(row=row, column=ci).fill = q_fill
            set_cell(ws7, row, 5, task_type, MUTED,   q_fill, CENTER)
            set_cell(ws7, row, 6, task_text, NORMAL,  q_fill, LEFT)
            ws7.cell(row=row, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws7.row_dimensions[row].height = 28
            row += 1

        # Spacer between products
        ws7.row_dimensions[row].height = 4
        for ci in range(1, 9):
            ws7.cell(row=row, column=ci).fill = hex_fill("f8fafc")
        row += 1

wb.save(OUTPUT_PATH)
print(f"✓ Excel saved → {OUTPUT_PATH}")
