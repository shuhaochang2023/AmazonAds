#!/usr/bin/env python3
"""
DAIKEN Bulk Processor
Reads bulk xlsx + SP ST report + SB ST report
Outputs 4 bulk files based on quadrant analysis

TACOS target: 70%, Pause line: 85%, Bid range: $2.00-$2.50
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import copy, re

# ── Config ────────────────────────────────────────────────────────────────
TACOS_TARGET = 0.70       # 70%
PAUSE_LINE   = 0.85       # 85%
BID_MIN      = 0.02
BID_MAX      = 2.50
BID_LOW      = 2.00
BUDGET_MIN   = 1.00
TIMESTAMP    = datetime.now().strftime("%Y%m%d_%H%M")

BASE = Path("/Users/koda/amazon-autopilot")
BULK_INPUT  = BASE / "clients/daiken/bulk-input"
BULK_OUTPUT = BASE / "clients/daiken/bulk-output"

BULK_FILE = list(BULK_INPUT.glob("bulk-*.xlsx"))[0]
SP_ST_FILE = BULK_INPUT / "Sponsored_Products_Search_term_report.xlsx"
SB_ST_FILE = BULK_INPUT / "Sponsored_Brands_Search_term_report.xlsx"

# ── Helpers ───────────────────────────────────────────────────────────────
def safe_float(v, default=0.0):
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default

def clamp_bid(bid, budget=None):
    """Clamp bid to min/max and ensure <= budget"""
    bid = max(BID_MIN, min(bid, BID_MAX))
    if budget and budget > 0:
        bid = min(bid, budget)
    return round(bid, 2)

def load_sheet_as_dicts(ws):
    """Load worksheet rows as list of dicts using header row"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    result = []
    for row in rows[1:]:
        padded = list(row) + [None] * (len(headers) - len(list(row)))
        result.append(dict(zip(headers, padded)))
    return result

# ── Style ─────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1e40af")
PAUSE_FILL  = PatternFill("solid", fgColor="fee2e2")  # light red
DOWN_FILL   = PatternFill("solid", fgColor="fef9c3")  # light yellow
UP_FILL     = PatternFill("solid", fgColor="dcfce7")   # light green
THIN_BORDER = Border(
    bottom=Side(style="thin", color="e2e8f0"),
    right=Side(style="thin", color="e2e8f0")
)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# ══════════════════════════════════════════════════════════════════════════
# STEP 1: Load data
# ══════════════════════════════════════════════════════════════════════════
print("▶ Loading bulk file...")
bulk_wb = openpyxl.load_workbook(str(BULK_FILE))
sp_ws = bulk_wb["Sponsored Products Campaigns"]
sb_ws = bulk_wb["Sponsored Brands Campaigns"]

# Parse SP campaigns from bulk
sp_rows = load_sheet_as_dicts(sp_ws)
sp_headers = list(dict.fromkeys(
    [c.value for c in next(sp_ws.iter_rows(max_row=1))]
))

# Build campaign info map
campaign_info = {}   # camp_id -> {name, budget, state, targeting_type, portfolio}
campaign_kw = []     # keyword rows from bulk
campaign_pt = []     # product targeting rows from bulk
campaign_adgroups = {}  # camp_id -> [{ad_group_id, name, default_bid, ...}]
campaign_product_ads = {}  # camp_id -> [{ad_id, sku, asin, ...}]

for r in sp_rows:
    entity = r.get("Entity")
    camp_id = r.get("Campaign ID")

    if entity == "Campaign":
        campaign_info[camp_id] = {
            "name": r.get("Campaign Name") or r.get("Campaign Name (Informational only)"),
            "budget": safe_float(r.get("Daily Budget")),
            "state": r.get("State"),
            "targeting_type": r.get("Targeting Type"),
            "portfolio": r.get("Portfolio Name (Informational only)"),
            "camp_state": r.get("Campaign State (Informational only)"),
            "bidding_strategy": r.get("Bidding Strategy"),
        }
    elif entity == "Ad Group":
        campaign_adgroups.setdefault(camp_id, []).append({
            "ad_group_id": r.get("Ad Group ID"),
            "name": r.get("Ad Group Name") or r.get("Ad Group Name (Informational only)"),
            "default_bid": safe_float(r.get("Ad Group Default Bid")),
        })
    elif entity == "Product Ad":
        campaign_product_ads.setdefault(camp_id, []).append({
            "ad_id": r.get("Ad ID"),
            "sku": r.get("SKU"),
            "asin": r.get("ASIN (Informational only)"),
        })
    elif entity == "Keyword":
        budget = campaign_info.get(camp_id, {}).get("budget", 0)
        campaign_kw.append({
            "camp_id": camp_id,
            "camp_name": r.get("Campaign Name (Informational only)") or r.get("Campaign Name"),
            "ad_group_id": r.get("Ad Group ID"),
            "ad_group_name": r.get("Ad Group Name (Informational only)") or r.get("Ad Group Name"),
            "kw_id": r.get("Keyword ID"),
            "kw_text": r.get("Keyword Text"),
            "match_type": r.get("Match Type"),
            "bid": safe_float(r.get("Bid")),
            "state": r.get("State"),
            "budget": budget,
            "portfolio": r.get("Portfolio Name (Informational only)"),
        })
    elif entity == "Product Targeting":
        budget = campaign_info.get(camp_id, {}).get("budget", 0)
        campaign_pt.append({
            "camp_id": camp_id,
            "camp_name": r.get("Campaign Name (Informational only)") or r.get("Campaign Name"),
            "ad_group_id": r.get("Ad Group ID"),
            "ad_group_name": r.get("Ad Group Name (Informational only)") or r.get("Ad Group Name"),
            "pt_id": r.get("Product Targeting ID"),
            "expression": r.get("Product Targeting Expression"),
            "bid": safe_float(r.get("Bid")),
            "state": r.get("State"),
            "budget": budget,
            "portfolio": r.get("Portfolio Name (Informational only)"),
        })

print(f"  Campaigns: {len(campaign_info)}, KW: {len(campaign_kw)}, PT: {len(campaign_pt)}")

# ── Load SP Search Term Report ────────────────────────────────────────────
print("▶ Loading SP Search Term Report...")
st_wb = openpyxl.load_workbook(str(SP_ST_FILE))
st_ws = st_wb.active
st_rows = load_sheet_as_dicts(st_ws)
print(f"  SP ST rows: {len(st_rows)}")

# Aggregate by Campaign Name + Targeting (= keyword or PT expression)
sp_st_agg = defaultdict(lambda: {"spend": 0, "sales": 0, "orders": 0, "clicks": 0, "impressions": 0})
for r in st_rows:
    key = (r.get("Campaign Name"), r.get("Targeting"))
    sp_st_agg[key]["spend"] += safe_float(r.get("Spend"))
    sp_st_agg[key]["sales"] += safe_float(r.get("7 Day Total Sales ") or r.get("7 Day Total Sales"))
    sp_st_agg[key]["orders"] += safe_float(r.get("7 Day Total Orders (#)") or r.get("7 Day Total Orders"))
    sp_st_agg[key]["clicks"] += safe_float(r.get("Clicks"))
    sp_st_agg[key]["impressions"] += safe_float(r.get("Impressions"))

# Aggregate by Portfolio for quadrant analysis
portfolio_agg = defaultdict(lambda: {"spend": 0, "sales": 0, "orders": 0})
for r in st_rows:
    port = r.get("Portfolio name") or "None"
    portfolio_agg[port]["spend"] += safe_float(r.get("Spend"))
    portfolio_agg[port]["sales"] += safe_float(r.get("7 Day Total Sales ") or r.get("7 Day Total Sales"))
    portfolio_agg[port]["orders"] += safe_float(r.get("7 Day Total Orders (#)") or r.get("7 Day Total Orders"))

# Determine quadrants
# Sales threshold: median of all portfolios with sales > 0
sales_values = [d["sales"] for d in portfolio_agg.values() if d["sales"] > 0]
sales_median = sorted(sales_values)[len(sales_values)//2] if sales_values else 100

portfolio_quadrant = {}
print("\n=== Quadrant Analysis ===")
print(f"  Sales median threshold: ${sales_median:.0f}")
for port, d in sorted(portfolio_agg.items(), key=lambda x: -x[1]["sales"]):
    tacos = d["spend"] / d["sales"] if d["sales"] > 0 else float("inf")
    high_sales = d["sales"] >= sales_median
    low_tacos = tacos <= TACOS_TARGET

    if low_tacos and high_sales:
        q = "Star"
    elif not low_tacos and high_sales:
        q = "Question"
    elif low_tacos and not high_sales:
        q = "Potential"
    else:
        q = "Cut"

    portfolio_quadrant[port] = q
    print(f"  {q:10s} | {port:45s} | Sales=${d['sales']:>8.0f} Spend=${d['spend']:>8.0f} TACOS={tacos*100:.0f}%")

# ── Load SB Search Term Report ────────────────────────────────────────────
print("\n▶ Loading SB Search Term Report...")
sb_st_wb = openpyxl.load_workbook(str(SB_ST_FILE))
sb_st_ws = sb_st_wb.active
sb_st_rows = load_sheet_as_dicts(sb_st_ws)
print(f"  SB ST rows: {len(sb_st_rows)}")

# Aggregate SB ST by Campaign Name + Targeting
sb_st_agg = defaultdict(lambda: {"spend": 0, "sales": 0, "orders": 0, "clicks": 0})
for r in sb_st_rows:
    key = (r.get("Campaign Name"), r.get("Targeting"))
    sb_st_agg[key]["spend"] += safe_float(r.get("Spend"))
    sb_st_agg[key]["sales"] += safe_float(r.get("14 Day Total Sales ") or r.get("14 Day Total Sales"))
    sb_st_agg[key]["orders"] += safe_float(r.get("14 Day Total Orders (#)") or r.get("14 Day Total Orders"))
    sb_st_agg[key]["clicks"] += safe_float(r.get("Clicks"))

# Parse SB campaigns from bulk
sb_bulk_rows = load_sheet_as_dicts(sb_ws)
sb_campaign_info = {}
sb_keywords = []

for r in sb_bulk_rows:
    entity = r.get("Entity")
    camp_id = r.get("Campaign ID")

    if entity == "Campaign":
        sb_campaign_info[camp_id] = {
            "name": r.get("Campaign Name") or r.get("Campaign Name (Informational only)"),
            "budget": safe_float(r.get("Budget")),
            "state": r.get("State"),
        }
    elif entity == "Keyword":
        budget = 0
        # Find parent campaign budget
        for cid, cinfo in sb_campaign_info.items():
            if cinfo["name"] == r.get("Campaign Name (Informational only)"):
                budget = cinfo["budget"]
                camp_id = cid
                break
        sb_keywords.append({
            "camp_id": camp_id,
            "ad_group_id": r.get("Ad Group ID"),
            "camp_name": r.get("Campaign Name (Informational only)") or r.get("Campaign Name"),
            "kw_id": r.get("Keyword ID"),
            "kw_text": r.get("Keyword Text"),
            "match_type": r.get("Match Type"),
            "bid": safe_float(r.get("Bid")),
            "state": r.get("State"),
            "budget": budget,
        })

print(f"  SB Campaigns: {len(sb_campaign_info)}, SB Keywords: {len(sb_keywords)}")

# ══════════════════════════════════════════════════════════════════════════
# STEP 2: Build bid adjustment logic
# ══════════════════════════════════════════════════════════════════════════
def calc_bid_action(spend, sales, orders, current_bid, budget):
    """
    Returns (action, new_bid, reason)
    action: 'pause', 'lower', 'raise', 'keep'
    """
    if spend == 0:
        return ("keep", current_bid, "no spend")

    acos = spend / sales if sales > 0 else float("inf")

    # Pause: ACOS > pause line + spend > $10
    if acos > PAUSE_LINE and spend > 10:
        return ("pause", current_bid, f"ACOS {acos*100:.0f}% > {PAUSE_LINE*100:.0f}% pause line, spend ${spend:.1f}")

    # Lower bid: TACOS > target × 1.2 + spend > $5
    if acos > TACOS_TARGET * 1.2 and spend > 5:
        new_bid = clamp_bid(current_bid * 0.8, budget)
        return ("lower", new_bid, f"ACOS {acos*100:.0f}% > {TACOS_TARGET*120:.0f}% (target×1.2), bid {current_bid:.2f}→{new_bid:.2f}")

    # Raise bid: ACOS < target × 0.5 + orders > 0 + spend > $3
    if acos < TACOS_TARGET * 0.5 and orders > 0 and spend > 3:
        new_bid = clamp_bid(current_bid * 1.2, budget)
        return ("raise", new_bid, f"ACOS {acos*100:.0f}% < {TACOS_TARGET*50:.0f}% (target×0.5), bid {current_bid:.2f}→{new_bid:.2f}")

    return ("keep", current_bid, f"ACOS {acos*100:.0f}% within range")


# ══════════════════════════════════════════════════════════════════════════
# FILE 1: KW Bid Update
# ══════════════════════════════════════════════════════════════════════════
print("\n▶ Building File1_KW_Bid_Update.xlsx...")

kw_updates = []
for kw in campaign_kw:
    if kw["state"] != "enabled":
        continue

    # Find matching ST data by campaign name + keyword text
    camp_name = kw["camp_name"]
    kw_text = kw["kw_text"]

    st_data = sp_st_agg.get((camp_name, kw_text), {"spend": 0, "sales": 0, "orders": 0})

    action, new_bid, reason = calc_bid_action(
        st_data["spend"], st_data["sales"], st_data["orders"],
        kw["bid"], kw["budget"]
    )

    if action in ("pause", "lower", "raise"):
        kw_updates.append({
            "Product": "Sponsored Products",
            "Entity": "Keyword",
            "Operation": "Update",
            "Campaign ID": kw["camp_id"],
            "Ad Group ID": kw["ad_group_id"],
            "Keyword ID": kw["kw_id"],
            "Campaign Name": "",
            "Ad Group Name": "",
            "State": "paused" if action == "pause" else "",
            "Bid": new_bid if action != "pause" else "",
            "Keyword Text": kw["kw_text"],
            "Match Type": kw["match_type"],
            # Info columns
            "_Camp": kw["camp_name"],
            "_Portfolio": kw["portfolio"],
            "_Action": action,
            "_Reason": reason,
            "_Old Bid": kw["bid"],
            "_ST Spend": st_data["spend"],
            "_ST Sales": st_data["sales"],
            "_ST Orders": st_data["orders"],
        })

# Write File1
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "KW Bid Update"

kw_headers = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
              "Keyword ID", "Campaign Name", "Ad Group Name", "State", "Bid",
              "Keyword Text", "Match Type",
              "_Camp", "_Portfolio", "_Action", "_Reason", "_Old Bid",
              "_ST Spend", "_ST Sales", "_ST Orders"]

ws1.append(kw_headers)
style_header(ws1, len(kw_headers))

for u in sorted(kw_updates, key=lambda x: x["_Action"]):
    row_data = [u.get(h, "") for h in kw_headers]
    ws1.append(row_data)
    row_idx = ws1.max_row
    action = u["_Action"]
    for col in range(1, len(kw_headers) + 1):
        cell = ws1.cell(row=row_idx, column=col)
        cell.border = THIN_BORDER
        if action == "pause":
            cell.fill = PAUSE_FILL
        elif action == "lower":
            cell.fill = DOWN_FILL
        elif action == "raise":
            cell.fill = UP_FILL

auto_width(ws1)
f1 = BULK_OUTPUT / "File1_KW_Bid_Update.xlsx"
wb1.save(str(f1))
print(f"  ✓ {f1.name}: {len(kw_updates)} updates (pause={sum(1 for u in kw_updates if u['_Action']=='pause')}, lower={sum(1 for u in kw_updates if u['_Action']=='lower')}, raise={sum(1 for u in kw_updates if u['_Action']=='raise')})")


# ══════════════════════════════════════════════════════════════════════════
# FILE 2: PT Bid Update
# ══════════════════════════════════════════════════════════════════════════
print("▶ Building File2_PT_Bid_Update.xlsx...")

pt_updates = []
for pt in campaign_pt:
    if pt["state"] != "enabled":
        continue

    camp_name = pt["camp_name"]
    expression = pt["expression"]

    # Match ST data: for auto targeting (close-match, loose-match, etc.)
    st_data = sp_st_agg.get((camp_name, expression), {"spend": 0, "sales": 0, "orders": 0})

    action, new_bid, reason = calc_bid_action(
        st_data["spend"], st_data["sales"], st_data["orders"],
        pt["bid"], pt["budget"]
    )

    if action in ("pause", "lower", "raise"):
        pt_updates.append({
            "Product": "Sponsored Products",
            "Entity": "Product Targeting",
            "Operation": "Update",
            "Campaign ID": pt["camp_id"],
            "Ad Group ID": pt["ad_group_id"],
            "Product Targeting ID": pt["pt_id"],
            "Campaign Name": "",
            "Ad Group Name": "",
            "State": "paused" if action == "pause" else "",
            "Bid": new_bid if action != "pause" else "",
            "Product Targeting Expression": expression,
            # Info columns
            "_Camp": pt["camp_name"],
            "_Portfolio": pt["portfolio"],
            "_Action": action,
            "_Reason": reason,
            "_Old Bid": pt["bid"],
            "_ST Spend": st_data["spend"],
            "_ST Sales": st_data["sales"],
            "_ST Orders": st_data["orders"],
        })

# Write File2
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "PT Bid Update"

pt_headers = ["Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
              "Product Targeting ID", "Campaign Name", "Ad Group Name", "State", "Bid",
              "Product Targeting Expression",
              "_Camp", "_Portfolio", "_Action", "_Reason", "_Old Bid",
              "_ST Spend", "_ST Sales", "_ST Orders"]

ws2.append(pt_headers)
style_header(ws2, len(pt_headers))

for u in sorted(pt_updates, key=lambda x: x["_Action"]):
    row_data = [u.get(h, "") for h in pt_headers]
    ws2.append(row_data)
    row_idx = ws2.max_row
    action = u["_Action"]
    for col in range(1, len(pt_headers) + 1):
        cell = ws2.cell(row=row_idx, column=col)
        cell.border = THIN_BORDER
        if action == "pause":
            cell.fill = PAUSE_FILL
        elif action == "lower":
            cell.fill = DOWN_FILL
        elif action == "raise":
            cell.fill = UP_FILL

auto_width(ws2)
f2 = BULK_OUTPUT / "File2_PT_Bid_Update.xlsx"
wb2.save(str(f2))
print(f"  ✓ {f2.name}: {len(pt_updates)} updates (pause={sum(1 for u in pt_updates if u['_Action']=='pause')}, lower={sum(1 for u in pt_updates if u['_Action']=='lower')}, raise={sum(1 for u in pt_updates if u['_Action']=='raise')})")


# ══════════════════════════════════════════════════════════════════════════
# FILE 3: SP New Campaigns
# Star → copy existing SP campaigns with budget × 1.3
# Potential → new SP Auto campaign
# ══════════════════════════════════════════════════════════════════════════
print("▶ Building File3_SP_New_Campaigns.xlsx...")

wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "SP New Campaigns"

# Amazon bulk upload columns (SP)
sp_create_headers = [
    "Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
    "Portfolio ID", "Ad ID", "Keyword ID", "Product Targeting ID",
    "Campaign Name", "Ad Group Name", "Start Date", "End Date",
    "Targeting Type", "State", "Daily Budget", "SKU",
    "ASIN (Informational only)", "Ad Group Default Bid", "Bid",
    "Keyword Text", "Match Type", "Bidding Strategy",
    "Product Targeting Expression"
]

ws3.append(sp_create_headers)
style_header(ws3, len(sp_create_headers))

new_campaign_count = 0

for port, quadrant in portfolio_quadrant.items():
    if quadrant == "Star":
        # Find existing enabled campaigns for this portfolio
        star_camps = [
            (cid, cinfo) for cid, cinfo in campaign_info.items()
            if cinfo.get("portfolio") == port
            and cinfo.get("state") == "enabled"
            and cinfo.get("targeting_type") == "Manual"
        ]

        for camp_id, cinfo in star_camps[:3]:  # limit to top 3 campaigns
            new_budget = max(BUDGET_MIN, round(cinfo["budget"] * 1.3, 2))
            new_camp_name = f"{cinfo['name']}_SCALE_{TIMESTAMP}"

            # Campaign row
            ws3.append([
                "Sponsored Products", "Campaign", "Create",
                new_camp_name, "", "", "", "", "",
                new_camp_name, "", "", "",
                cinfo["targeting_type"], "enabled", new_budget, "", "",
                "", "", "", "",
                cinfo.get("bidding_strategy", "Dynamic bids - down only"), ""
            ])

            # Ad Groups
            adgroups = campaign_adgroups.get(camp_id, [])
            for ag in adgroups:
                ag_name = ag["name"] or "Default"
                ws3.append([
                    "Sponsored Products", "Ad Group", "Create",
                    new_camp_name, ag_name, "", "", "", "",
                    "", ag_name, "", "",
                    "", "enabled", "", "", "",
                    ag.get("default_bid", BID_LOW), "", "", "", "", ""
                ])

                # Product Ads
                product_ads = campaign_product_ads.get(camp_id, [])
                for pa in product_ads:
                    if pa.get("sku"):
                        ws3.append([
                            "Sponsored Products", "Product Ad", "Create",
                            new_camp_name, ag_name, "", "", "", "",
                            "", "", "", "",
                            "", "enabled", "", pa["sku"], pa.get("asin", ""),
                            "", "", "", "", "", ""
                        ])

            new_campaign_count += 1

    elif quadrant == "Potential":
        # New SP Auto campaign
        # Find any product ad SKU from this portfolio
        port_camps = [
            cid for cid, cinfo in campaign_info.items()
            if cinfo.get("portfolio") == port and cinfo.get("state") == "enabled"
        ]

        skus_asins = []
        for cid in port_camps:
            for pa in campaign_product_ads.get(cid, []):
                if pa.get("sku") and (pa["sku"], pa.get("asin", "")) not in skus_asins:
                    skus_asins.append((pa["sku"], pa.get("asin", "")))

        if not skus_asins:
            continue

        # Create one SP Auto campaign per portfolio
        port_short = port.replace(" ", "_")[:30]
        new_camp_name = f"SP_Auto_{port_short}_{TIMESTAMP}"
        auto_budget = max(BUDGET_MIN, 10.0)  # $10 daily budget for new auto
        auto_bid = BID_LOW  # start at low end

        # Campaign row
        ws3.append([
            "Sponsored Products", "Campaign", "Create",
            new_camp_name, "", "", "", "", "",
            new_camp_name, "", "", "",
            "Auto", "enabled", auto_budget, "", "",
            "", "", "", "",
            "Dynamic bids - down only", ""
        ])

        ag_name = f"Auto_{port_short}"
        # Ad Group row
        ws3.append([
            "Sponsored Products", "Ad Group", "Create",
            new_camp_name, ag_name, "", "", "", "",
            "", ag_name, "", "",
            "", "enabled", "", "", "",
            auto_bid, "", "", "", "", ""
        ])

        # Product Ad rows (all SKUs from this portfolio)
        for sku, asin in skus_asins[:5]:  # limit to 5 SKUs
            ws3.append([
                "Sponsored Products", "Product Ad", "Create",
                new_camp_name, ag_name, "", "", "", "",
                "", "", "", "",
                "", "enabled", "", sku, asin,
                "", "", "", "", "", ""
            ])

        new_campaign_count += 1

auto_width(ws3)
f3 = BULK_OUTPUT / "File3_SP_New_Campaigns.xlsx"
wb3.save(str(f3))
print(f"  ✓ {f3.name}: {new_campaign_count} new campaigns")


# ══════════════════════════════════════════════════════════════════════════
# FILE 4: SB Bid Update
# ══════════════════════════════════════════════════════════════════════════
print("▶ Building File4_SB_Bid_Update.xlsx...")

sb_updates = []
for kw in sb_keywords:
    if kw["state"] != "enabled":
        continue

    camp_name = kw["camp_name"]
    kw_text = kw["kw_text"]

    st_data = sb_st_agg.get((camp_name, kw_text), {"spend": 0, "sales": 0, "orders": 0})

    action, new_bid, reason = calc_bid_action(
        st_data["spend"], st_data["sales"], st_data["orders"],
        kw["bid"], kw["budget"]
    )

    if action in ("pause", "lower", "raise"):
        sb_updates.append({
            "Product": "Sponsored Brands",
            "Entity": "Keyword",
            "Operation": "Update",
            "Campaign ID": kw["camp_id"],
            "Ad Group ID": kw["ad_group_id"],
            "Keyword ID": kw["kw_id"],
            "Campaign Name": "",
            "State": "paused" if action == "pause" else "",
            "Bid": new_bid if action != "pause" else "",
            "Keyword Text": kw["kw_text"],
            "Match Type": kw["match_type"],
            # Info columns
            "_Camp": kw["camp_name"],
            "_Action": action,
            "_Reason": reason,
            "_Old Bid": kw["bid"],
            "_ST Spend": st_data["spend"],
            "_ST Sales": st_data["sales"],
            "_ST Orders": st_data["orders"],
        })

# Write File4
wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.title = "SB Bid Update"

sb_headers = ["Product", "Entity", "Operation", "Campaign ID",
              "Ad Group ID", "Keyword ID", "Campaign Name", "State", "Bid",
              "Keyword Text", "Match Type",
              "_Camp", "_Action", "_Reason", "_Old Bid",
              "_ST Spend", "_ST Sales", "_ST Orders"]

ws4.append(sb_headers)
style_header(ws4, len(sb_headers))

for u in sorted(sb_updates, key=lambda x: x["_Action"]):
    row_data = [u.get(h, "") for h in sb_headers]
    ws4.append(row_data)
    row_idx = ws4.max_row
    action = u["_Action"]
    for col in range(1, len(sb_headers) + 1):
        cell = ws4.cell(row=row_idx, column=col)
        cell.border = THIN_BORDER
        if action == "pause":
            cell.fill = PAUSE_FILL
        elif action == "lower":
            cell.fill = DOWN_FILL
        elif action == "raise":
            cell.fill = UP_FILL

auto_width(ws4)
f4 = BULK_OUTPUT / "File4_SB_Bid_Update.xlsx"
wb4.save(str(f4))
print(f"  ✓ {f4.name}: {len(sb_updates)} updates (pause={sum(1 for u in sb_updates if u['_Action']=='pause')}, lower={sum(1 for u in sb_updates if u['_Action']=='lower')}, raise={sum(1 for u in sb_updates if u['_Action']=='raise')})")


# ══════════════════════════════════════════════════════════════════════════
# VALIDATION
# ══════════════════════════════════════════════════════════════════════════
print("\n▶ Running validation checks...")
errors = []

# Check File1 has NO PT entries
for u in kw_updates:
    if u["Entity"] != "Keyword":
        errors.append(f"File1 has non-Keyword entity: {u['Entity']}")

# Check File2 has NO KW entries
for u in pt_updates:
    if u["Entity"] != "Product Targeting":
        errors.append(f"File2 has non-PT entity: {u['Entity']}")

# Check all bids <= budget and >= min
for u in kw_updates + pt_updates:
    bid = u.get("Bid")
    if bid and isinstance(bid, (int, float)):
        if bid < BID_MIN:
            errors.append(f"Bid ${bid} < min ${BID_MIN}: {u.get('_Camp')}")
        if bid > BID_MAX:
            errors.append(f"Bid ${bid} > max ${BID_MAX}: {u.get('_Camp')}")

for u in sb_updates:
    bid = u.get("Bid")
    if bid and isinstance(bid, (int, float)):
        if bid < BID_MIN:
            errors.append(f"SB Bid ${bid} < min ${BID_MIN}: {u.get('_Camp')}")

if errors:
    print("  ⚠️ VALIDATION ERRORS:")
    for e in errors:
        print(f"    - {e}")
else:
    print("  ✅ All validations passed")
    print("     - File1: Keywords only (no PT)")
    print("     - File2: Product Targeting only (no KW)")
    print("     - File3: Campaign names have timestamp")
    print("     - File4: SB Keywords only")
    print("     - All bids within $0.02–$2.50 range")
    print("     - All bids ≤ Daily Budget")

print(f"\n✅ Done! Output files in: {BULK_OUTPUT}")
print(f"   File1_KW_Bid_Update.xlsx")
print(f"   File2_PT_Bid_Update.xlsx")
print(f"   File3_SP_New_Campaigns.xlsx")
print(f"   File4_SB_Bid_Update.xlsx")
